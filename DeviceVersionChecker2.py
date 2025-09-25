#!/usr/bin/env python3
"""
Device Version Checker - SSH into network devices and generate version compliance report

This script connects to network devices via SSH using netmiko, collects 'show version'
information, compares device versions against known good versions, and generates
an Excel report with conditional formatting.

Author: SuperDan Environment
Date: 2025-09-25
"""

from netmiko import ConnectHandler
from tools import getScriptName, setupLogging, get_netmiko_creds, save_file_and_set_permissions
import logging
import csv
import re
from datetime import datetime
import sys
import os

# Excel libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl library not found. Please install it with: pip install openpyxl")
    sys.exit(1)


class DeviceVersionChecker:
    """Class to handle device version checking and reporting"""
    
    def __init__(self):
        self.script_name = getScriptName()
        setupLogging()
        self.logger = logging.getLogger(__name__)
        
        # Get credentials
        try:
            self.netmiko_user, self.passwd, self.enable = get_netmiko_creds()
        except:
            # Fallback to tools.py global variables if function format is different
            from tools import netmikoUser, passwd, enable
            self.netmiko_user = netmikoUser
            self.passwd = passwd
            self.enable = enable
        
        self.devices = []
        self.known_good_versions = {}
        self.results = []
        
        # Excel styling
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.header_font = Font(bold=True)
        self.center_alignment = Alignment(horizontal="center", vertical="center")

    def load_devices(self, csv_file="devices.csv"):
        """Load device information from CSV file"""
        if not os.path.exists(csv_file):
            self.logger.error(f"Device CSV file not found: {csv_file}")
            self.logger.info("Please create devices.csv with columns: devicename,ipAddr,deviceGroup,deviceType")
            return False
        
        try:
            with open(csv_file, 'r') as file:
                reader = csv.DictReader(file)
                all_devices = list(reader)
            
            # Filter to only include cisco_ios and cisco_ios_xe devices
            supported_types = ['cisco_ios', 'cisco_ios_xe']
            self.devices = [device for device in all_devices 
                          if device.get('deviceType', '').strip() in supported_types]
            
            skipped_count = len(all_devices) - len(self.devices)
            
            self.logger.info(f"Loaded {len(self.devices)} cisco_ios/cisco_ios_xe devices from {csv_file}")
            if skipped_count > 0:
                self.logger.info(f"Skipped {skipped_count} devices with unsupported device types")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading devices from {csv_file}: {e}")
            return False

    def load_known_good_versions(self, csv_file="known_good_versions.csv"):
        """Load known good versions from CSV file"""
        if not os.path.exists(csv_file):
            self.logger.error(f"Known good versions file not found: {csv_file}")
            self.logger.info("Please create known_good_versions.csv with columns: device_type,version")
            return False
        
        try:
            # Initialize as dictionary with lists to handle multiple versions per device type
            self.known_good_versions = {}
            
            with open(csv_file, 'r') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    device_type = row['device_type'].strip()
                    version = row['version'].strip()
                    
                    # Create list if device type doesn't exist
                    if device_type not in self.known_good_versions:
                        self.known_good_versions[device_type] = []
                    
                    # Add version to the list
                    self.known_good_versions[device_type].append(version)
            
            total_versions = sum(len(versions) for versions in self.known_good_versions.values())
            self.logger.info(f"Loaded {total_versions} known good versions for {len(self.known_good_versions)} device types")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading known good versions from {csv_file}: {e}")
            return False

    def detect_actual_device_type(self, output):
        """Detect if device is actually IOS XE based on show version output"""
        if re.search(r'Cisco IOS XE Software', output, re.IGNORECASE):
            return 'cisco_ios_xe'
        else:
            return 'cisco_ios'
    
    def extract_version_from_output(self, output, device_type):
        """Extract version number from show version output"""
        # Detect actual device type from output
        actual_device_type = self.detect_actual_device_type(output)
        
        version_patterns = {
            'cisco_ios': [
                r'Cisco IOS Software.*Version\s+([0-9]+\.[0-9]+(?:\.[0-9]+)*[A-Z0-9]*)',
                r'Version\s+([0-9]+\.[0-9]+(?:\.[0-9]+)*[A-Z0-9]*)',
            ],
            'cisco_ios_xe': [
                r'Cisco IOS XE Software.*Version\s+([0-9]+\.[0-9]+\.[0-9]+[A-Z0-9]*)',
                r'Version\s+([0-9]+\.[0-9]+\.[0-9]+[A-Z0-9]*)',
            ]
        }
        
        # Use actual device type for pattern matching
        patterns = version_patterns.get(actual_device_type, version_patterns['cisco_ios'])
        
        for pattern in patterns:
            match = re.search(pattern, output, re.IGNORECASE)
            if match:
                version = match.group(1)
                # Clean up version string (remove trailing letters/codes)
                clean_version = re.sub(r'[A-Z]+$', '', version)
                return clean_version, actual_device_type
        
        self.logger.warning(f"Could not extract version from output")
        return "Unknown", actual_device_type

    def find_matching_known_good_version(self, current_version, actual_device_type):
        """Find the appropriate known good version based on version train"""
        if current_version == "Unknown":
            return "Unknown"
            
        try:
            current_parts = [int(x) for x in current_version.split('.')]
            current_major_minor = f"{current_parts[0]}.{current_parts[1]}"
            
            # Get versions for this device type
            device_versions = self.known_good_versions.get(actual_device_type, [])
            
            if not device_versions:
                return "Not Configured"
            
            # First try to find exact version train match (major.minor)
            for known_version in device_versions:
                try:
                    known_parts = [int(x) for x in known_version.split('.')]
                    known_major_minor = f"{known_parts[0]}.{known_parts[1]}"
                    
                    # Check if it's the same version train (major.minor)
                    if current_major_minor == known_major_minor:
                        return known_version
                        
                except (ValueError, IndexError):
                    continue
            
            # If no exact train match, return the first available version for this device type
            # This provides a reference point even if not same train
            return device_versions[0]
            
        except (ValueError, IndexError) as e:
            self.logger.warning(f"Error finding matching version for {current_version}: {e}")
            return "Not Configured"
    
    def compare_versions(self, current_version, known_good_version):
        """Compare current version with known good version"""
        if current_version == "Unknown" or known_good_version in ["Unknown", "Not Configured"]:
            return "Unknown" if current_version == "Unknown" else "No Reference"
        
        try:
            # Split versions into components for comparison
            current_parts = [int(x) for x in current_version.split('.')]
            good_parts = [int(x) for x in known_good_version.split('.')]
            
            # Pad shorter version with zeros
            max_len = max(len(current_parts), len(good_parts))
            current_parts.extend([0] * (max_len - len(current_parts)))
            good_parts.extend([0] * (max_len - len(good_parts)))
            
            # Compare version parts
            if current_parts < good_parts:
                return "Below"
            elif current_parts == good_parts:
                return "Match"
            else:
                return "Above"
                
        except ValueError as e:
            self.logger.warning(f"Error comparing versions {current_version} vs {known_good_version}: {e}")
            return "Error"

    def connect_and_get_version(self, device):
        """Connect to device and get version information"""
        devicename = device['devicename']
        ip_addr = device['ipAddr']
        device_group = device.get('deviceGroup', 'Unknown')
        device_type = device['deviceType']
        
        self.logger.info(f"Connecting to {devicename} ({ip_addr})")
        
        try:
            # Create connection
            connection_params = {
                'device_type': device_type,
                'host': ip_addr,
                'username': self.netmiko_user,
                'password': self.passwd,
                'secret': self.enable,
                'timeout': 60,
                'session_timeout': 60,
                'auth_timeout': 60,
                'global_delay_factor': 2,
                'fast_cli': False,
            }
            
            conn = ConnectHandler(**connection_params)
            
            # Enter enable mode if needed
            if hasattr(conn, 'enable'):
                conn.enable()
            
            # Get version information
            version_output = conn.send_command('show version', use_textfsm=False)
            
            # Extract version and detect actual device type
            current_version, actual_device_type = self.extract_version_from_output(version_output, device_type)
            
            # Find matching known good version based on version train
            known_good = self.find_matching_known_good_version(current_version, actual_device_type)
            
            # Compare versions
            comparison = self.compare_versions(current_version, known_good)
            
            # Store results
            result = {
                'devicename': devicename,
                'ip_addr': ip_addr,
                'device_group': device_group,
                'device_type': actual_device_type,  # Use detected device type
                'csv_device_type': device_type,     # Original CSV device type
                'current_version': current_version,
                'known_good_version': known_good,
                'comparison': comparison,
                'status': 'Success',
                'notes': ''
            }
            
            conn.disconnect()
            self.logger.info(f"Successfully collected version from {devicename}: {current_version}")
            
            return result
            
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"Failed to connect to {devicename} ({ip_addr}): {error_msg}")
            
            return {
                'devicename': devicename,
                'ip_addr': ip_addr,
                'device_group': device_group,
                'device_type': device_type,
                'csv_device_type': device_type,
                'current_version': 'Connection Failed',
                'known_good_version': 'Connection Failed',
                'comparison': 'Error',
                'status': 'Failed',
                'notes': error_msg
            }

    def collect_all_versions(self):
        """Collect version information from all devices"""
        self.logger.info(f"Starting version collection from {len(self.devices)} devices")
        
        for device in self.devices:
            result = self.connect_and_get_version(device)
            self.results.append(result)
        
        self.logger.info("Version collection completed")

    def generate_excel_report(self, filename=None):
        """Generate Excel report with conditional formatting"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.script_name}_report_{timestamp}.xlsx"
        
        self.logger.info(f"Generating Excel report: {filename}")
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Device Version Report"
        
        # Define headers
        headers = [
            'Device Name',
            'IP Address',
            'Device Group',
            'Detected Type',
            'CSV Type',
            'Current Version',
            'Known Good Version',
            'Comparison',
            'Status',
            'Notes'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
        
        # Add data rows
        for row_idx, result in enumerate(self.results, 2):
            # Add data
            ws.cell(row=row_idx, column=1, value=result['devicename'])
            ws.cell(row=row_idx, column=2, value=result['ip_addr'])
            ws.cell(row=row_idx, column=3, value=result['device_group'])
            ws.cell(row=row_idx, column=4, value=result['device_type'])
            ws.cell(row=row_idx, column=5, value=result['csv_device_type'])
            ws.cell(row=row_idx, column=6, value=result['current_version'])
            ws.cell(row=row_idx, column=7, value=result['known_good_version'])
            ws.cell(row=row_idx, column=8, value=result['comparison'])
            ws.cell(row=row_idx, column=9, value=result['status'])
            ws.cell(row=row_idx, column=10, value=result['notes'])
            
            # Apply yellow highlighting for devices with versions below known good
            if result['comparison'] == 'Below':
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = self.yellow_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filename)
        
        # Set file permissions
        save_file_and_set_permissions(filename, show_info=True)
        
        self.logger.info(f"Excel report generated: {filename}")
        return filename

    def print_summary(self):
        """Print summary statistics"""
        total_devices = len(self.results)
        successful_connections = len([r for r in self.results if r['status'] == 'Success'])
        below_good_version = len([r for r in self.results if r['comparison'] == 'Below'])
        
        # Count device types processed
        ios_devices = len([r for r in self.results if r['device_type'] == 'cisco_ios'])
        ios_xe_devices = len([r for r in self.results if r['device_type'] == 'cisco_ios_xe'])
        
        print(f"\n=== Device Version Check Summary ===")
        print(f"Total cisco_ios/cisco_ios_xe devices processed: {total_devices}")
        print(f"  - Detected as IOS: {ios_devices}")
        print(f"  - Detected as IOS XE: {ios_xe_devices}")
        print(f"Successful connections: {successful_connections}")
        print(f"Failed connections: {total_devices - successful_connections}")
        print(f"Devices below known good version: {below_good_version}")
        
        if below_good_version > 0:
            print(f"\nDevices requiring updates:")
            for result in self.results:
                if result['comparison'] == 'Below':
                    print(f"  - {result['devicename']}: {result['current_version']} (should be {result['known_good_version']})")


def main():
    """Main function"""
    checker = DeviceVersionChecker()
    
    # Load configuration files
    if not checker.load_devices():
        return 1
    
    if not checker.load_known_good_versions():
        return 1
    
    # Collect version information
    checker.collect_all_versions()
    
    # Generate Excel report
    report_filename = checker.generate_excel_report()
    
    # Print summary
    checker.print_summary()
    
    print(f"\nReport saved as: {report_filename}")
    print("Devices with versions below known good are highlighted in yellow.")
    
    return 0


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
