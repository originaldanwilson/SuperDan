from netmiko import ConnectHandler
from GetCreds import get_netmiko_creds
from tools import getScriptName, setupLogging
from openpyxl import Workbook
import logging
import re
from collections import defaultdict

def get_fex_label(interface_name):
    match = re.match(r"Eth(\d+)", interface_name)
    if match:
        slot = int(match.group(1))
        return str(slot) if slot >= 100 else "native"
    return "unknown"

def collect_connected_interfaces(hostname, netmikoUser, passwd, sheet_map, summary_tracker):
    try:
        conn = ConnectHandler(
            device_type="cisco_nxos",
            host=hostname,
            username=netmikoUser,
            password=passwd
        )
        output = conn.send_command("show interface status", use_textfsm=False)
        conn.disconnect()

        for line in output.splitlines():
            if "connected" in line.lower():
                parts = line.split()
                if len(parts) >= 2:
                    interface = parts[0]
                    status = parts[1]
                    description = " ".join(parts[6:]) if len(parts) > 6 else ""
                    label = get_fex_label(interface)

                    ws = sheet_map[label]
                    ws.append([hostname, interface, status, description])
                    summary_tracker[hostname][label] += 1

    except Exception as e:
        logging.error(f"Failed to process {hostname}: {e}")

def init_sheet(wb, label):
    ws = wb.create_sheet(title=label)
    ws.append(["switch", "interface", "status", "description"])
    ws.freeze_panes = "A2"
    return ws

def auto_width_all(wb):
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

def write_summary(wb, summary_tracker, all_labels):
    ws = wb.create_sheet(title="summary", index=0)
    header = ["switch"] + all_labels
    ws.append(header)

    for hostname in sorted(summary_tracker.keys()):
        row = [hostname]
        for label in all_labels:
            row.append(summary_tracker[hostname].get(label, 0))
        ws.append(row)

    ws.freeze_panes = "B2"

def main():
    scriptName = getScriptName()
    setupLogging(scriptName)
    logging.info(f"{scriptName} started")

    netmikoUser, passwd, enable = get_netmiko_creds()

    switches = [
        "sw1", "sw2", "sw3", "sw4", "sw5", "sw6", "sw7",
        "sw8", "sw9", "sw10", "sw11", "sw12", "sw13", "sw14",
        "sw15", "sw16", "sw17", "sw18", "sw19", "sw20",
        "sw21", "sw22", "sw23", "sw24", "sw25"
    ]

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    sheet_map = {}
    all_labels = ["native"] + [str(f) for f in range(100, 200)]
    for label in all_labels:
        sheet_map[label] = init_sheet(wb, label)

    summary_tracker = defaultdict(lambda: defaultdict(int))

    for hostname in switches:
        collect_connected_interfaces(hostname, netmikoUser, passwd, sheet_map, summary_tracker)

    write_summary(wb, summary_tracker, all_labels)
    auto_width_all(wb)

    outname = f"{scriptName}.xlsx"
    wb.save(outname)
    logging.info(f"Saved spreadsheet: {outname}")

if __name__ == "__main__":
    main()
