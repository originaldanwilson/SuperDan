import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from netmiko import ConnectHandler
from GetCreds import get_netmiko_creds
from tools import getScriptName, setupLogging, logScriptStart, outputFile

def extract_pid_sn_lines(output):
    results = []
    current_iface = None
    pid = "UNKNOWN"
    sn = "UNKNOWN"

    for line in output.splitlines():
        line = line.strip()
        if line.startswith("Ethernet"):
            current_iface = line.split()[0]
        elif "product id is" in line and current_iface:
            parts = line.split()
            if len(parts) >= 4:
                pid = parts[3]
        elif "serial number is" in line and current_iface:
            parts = line.split()
            sn = parts[-1] if len(parts) >= 1 else "UNKNOWN"
            results.append((current_iface, pid, sn))
            current_iface = None
            pid = "UNKNOWN"
            sn = "UNKNOWN"
    return results

def main():
    scriptName = getScriptName()
    logger = setupLogging(scriptName)
    logScriptStart(logger, scriptName)
    outputPath = outputFile()

    switchList = [
        "switch1.example.com",
        "switch2.example.com"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFP Details"
    ws.append(["Switch", "Interface", "Product ID", "Serial Number"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    netmikoUser, passwd, enable = get_netmiko_creds()

    for switch in switchList:
        logger.info(f"Connecting to {switch}")
        device = {
            "device_type": "cisco_nxos",
            "host": switch,
            "username": netmikoUser,
            "password": passwd,
            "secret": enable,
        }

        try:
            conn = ConnectHandler(**device)
            conn.enable()
            output = conn.send_command("show interface transceiver detail | include Ethernet|product|serial")
            conn.disconnect()

            parsed = extract_pid_sn_lines(output)
            for iface, pid, sn in parsed:
                ws.append([switch, iface, pid, sn])

        except Exception as e:
            logger.error(f"{switch} failed: {e}")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(outputPath)
    logger.info(f"Saved spreadsheet to {outputPath}")

if __name__ == "__main__":
    main()
