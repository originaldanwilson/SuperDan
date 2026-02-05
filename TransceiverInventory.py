#!/usr/bin/env python3
"""
Transceiver Inventory - Collect SFP/transceiver details from network devices

This script connects to network devices via SSH using netmiko with threading,
collects 'show interfaces transceiver' information, parses it using TextFSM
or regex fallback, and generates an Excel report with a summary page.

Author: SuperDan Environment
"""

from netmiko import ConnectHandler
from tools import getScriptName, setupLogging, get_netmiko_creds, save_file_and_set_permissions
import logging
import csv
import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter
import threading
import sys
import os

# Excel libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl library not found. Please install it with: pip install openpyxl")
    sys.exit(1)

# TextFSM support (optional - will use regex fallback if not available)
try:
    import textfsm
    TEXTFSM_AVAILABLE = True
except ImportError:
    TEXTFSM_AVAILABLE = False


class TransceiverInventory:
    """Class to handle transceiver inventory collection and reporting"""
    
    # Maximum concurrent connections
    MAX_WORKERS = 20
    
    def __init__(self, csv_file="freedevices.csv"):
        self.script_name = getScriptName()
        setupLogging(self.script_name)
        self.logger = logging.getLogger(__name__)
        self.csv_file = csv_file
        
        # Get credentials
        try:
            self.netmiko_user, self.passwd, self.enable = get_netmiko_creds()
        except:
            from tools import netmikoUser, passwd, enable
            self.netmiko_user = netmikoUser
            self.passwd = passwd
            self.enable = enable
        
        self.devices = []
        self.results = []  # List of transceiver records
        self.device_errors = []  # Track failed devices
        self.results_lock = threading.Lock()
        
        # Excel styling
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.center_alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
        self.left_alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
        self.summary_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    def load_devices(self):
        """Load device information from CSV file"""
        if not os.path.exists(self.csv_file):
            self.logger.error(f"Device CSV file not found: {self.csv_file}")
            self.logger.info("Please create freedevices.csv with columns: devicename,ipaddr,description,devicetype")
            return False
        
        try:
            with open(self.csv_file, 'r', encoding='utf-8-sig') as file:
                reader = csv.DictReader(file)
                all_devices = list(reader)
            
            # Normalize column names (handle various formats)
            normalized_devices = []
            for device in all_devices:
                normalized = {}
                for key, value in device.items():
                    norm_key = key.strip().lower().replace(' ', '_').replace('addr', 'addr')
                    # Handle common variations
                    if 'device' in norm_key and 'name' in norm_key:
                        norm_key = 'devicename'
                    elif norm_key in ['ip', 'ip_addr', 'ipaddr', 'ip_address']:
                        norm_key = 'ipaddr'
                    elif norm_key in ['os', 'operating_system', 'device_type', 'devicetype']:
                        norm_key = 'devicetype'
                    elif norm_key in ['desc', 'description']:
                        norm_key = 'description'
                    normalized[norm_key] = value.strip() if value else ''
                normalized_devices.append(normalized)
            
            # Filter to only include cisco_ios and cisco_nxos devices
            supported_types = ['cisco_ios', 'cisco_nxos']
            self.devices = [d for d in normalized_devices 
                          if d.get('devicetype', '').lower() in supported_types]
            
            skipped_count = len(normalized_devices) - len(self.devices)
            
            self.logger.info(f"Loaded {len(self.devices)} cisco_ios/cisco_nxos devices from {self.csv_file}")
            if skipped_count > 0:
                self.logger.info(f"Skipped {skipped_count} devices with unsupported device types")
            
            return len(self.devices) > 0
            
        except Exception as e:
            self.logger.error(f"Error loading devices from {self.csv_file}: {e}")
            return False

    def parse_nxos_transceiver(self, output, devicename):
        """Parse NX-OS show interface transceiver output"""
        transceivers = []
        
        # NX-OS format parsing - each interface block
        current_interface = None
        current_data = {}
        
        for line in output.splitlines():
            line = line.strip()
            
            # Interface line (e.g., "Ethernet1/1")
            if re.match(r'^(Ethernet\d+/\d+(/\d+)?|port-channel\d+)', line, re.IGNORECASE):
                # Save previous interface if we have data
                if current_interface and current_data:
                    transceivers.append(current_data)
                
                current_interface = line.split()[0]
                current_data = {
                    'devicename': devicename,
                    'interface': current_interface,
                    'cisco_part_number': '',
                    'serial_number': '',
                    'type': '',
                    'cisco_product_id': '',
                    'name': ''
                }
            
            # Parse various fields
            elif current_interface:
                # Type/transceiver type
                if 'transceiver is' in line.lower():
                    match = re.search(r'transceiver is\s+(.+)', line, re.IGNORECASE)
                    if match:
                        current_data['type'] = match.group(1).strip()
                
                # Cisco part number
                elif 'cisco part number' in line.lower():
                    match = re.search(r'cisco part number[:\s]+(\S+)', line, re.IGNORECASE)
                    if match:
                        current_data['cisco_part_number'] = match.group(1)
                
                # Cisco product id
                elif 'cisco product id' in line.lower() or 'product id' in line.lower():
                    match = re.search(r'(?:cisco )?product id[:\s]+(\S+)', line, re.IGNORECASE)
                    if match:
                        current_data['cisco_product_id'] = match.group(1)
                
                # Serial number
                elif 'serial number' in line.lower():
                    match = re.search(r'serial number[:\s]+(\S+)', line, re.IGNORECASE)
                    if match:
                        current_data['serial_number'] = match.group(1)
                
                # Name/description
                elif 'name' in line.lower() and 'is' in line.lower():
                    match = re.search(r'name\s+is\s+(.+)', line, re.IGNORECASE)
                    if match:
                        current_data['name'] = match.group(1).strip()
        
        # Don't forget the last interface
        if current_interface and current_data:
            transceivers.append(current_data)
        
        # Filter out entries without meaningful data
        transceivers = [t for t in transceivers if t.get('cisco_part_number') or t.get('serial_number')]
        
        return transceivers

    def parse_ios_transceiver(self, output, devicename):
        """Parse IOS show interface transceiver output"""
        transceivers = []
        
        current_interface = None
        current_data = {}
        
        for line in output.splitlines():
            line = line.strip()
            
            # Interface patterns for IOS
            interface_match = re.match(r'^(Gi\S+|Te\S+|Fa\S+|Eth\S+|TwoGig\S+|TenGig\S+|TwentyFiveGig\S+|FortyGig\S+|HundredGig\S+)', line, re.IGNORECASE)
            if interface_match:
                if current_interface and current_data:
                    transceivers.append(current_data)
                
                current_interface = interface_match.group(1)
                current_data = {
                    'devicename': devicename,
                    'interface': current_interface,
                    'cisco_part_number': '',
                    'serial_number': '',
                    'type': '',
                    'cisco_product_id': '',
                    'name': ''
                }
            
            elif current_interface:
                # Parse various fields - IOS format
                if 'transceiver type' in line.lower():
                    match = re.search(r'transceiver type[:\s]+(.+)', line, re.IGNORECASE)
                    if match:
                        current_data['type'] = match.group(1).strip()
                
                elif 'part number' in line.lower():
                    match = re.search(r'part number[:\s]+(\S+)', line, re.IGNORECASE)
                    if match:
                        current_data['cisco_part_number'] = match.group(1)
                
                elif 'product id' in line.lower():
                    match = re.search(r'product id[:\s]+(\S+)', line, re.IGNORECASE)
                    if match:
                        current_data['cisco_product_id'] = match.group(1)
                
                elif 'serial' in line.lower():
                    match = re.search(r'serial[^:]*[:\s]+(\S+)', line, re.IGNORECASE)
                    if match:
                        current_data['serial_number'] = match.group(1)
                
                elif 'vendor name' in line.lower() or ('name' in line.lower() and 'is' in line.lower()):
                    match = re.search(r'(?:vendor )?name[:\s]+(.+)', line, re.IGNORECASE)
                    if match:
                        current_data['name'] = match.group(1).strip()
        
        # Don't forget the last interface
        if current_interface and current_data:
            transceivers.append(current_data)
        
        # Filter out entries without meaningful data
        transceivers = [t for t in transceivers if t.get('cisco_part_number') or t.get('serial_number')]
        
        return transceivers

    def connect_and_collect(self, device):
        """Connect to device and collect transceiver information"""
        devicename = device.get('devicename', 'Unknown')
        ip_addr = device.get('ipaddr', '')
        description = device.get('description', '')
        device_type = device.get('devicetype', 'cisco_ios').lower()
        
        self.logger.info(f"Connecting to {devicename} ({ip_addr})")
        
        transceivers = []
        
        try:
            connection_params = {
                'device_type': device_type,
                'host': ip_addr,
                'username': self.netmiko_user,
                'password': self.passwd,
                'secret': self.enable,
                'timeout': 60,
                'session_timeout': 120,
                'auth_timeout': 60,
                'global_delay_factor': 2,
                'fast_cli': False,
            }
            
            conn = ConnectHandler(**connection_params)
            
            if hasattr(conn, 'enable'):
                try:
                    conn.enable()
                except:
                    pass  # Some devices don't need enable
            
            # Run show interface transceiver command
            if device_type == 'cisco_nxos':
                # NX-OS command - get detail for more info
                output = conn.send_command('show interface transceiver detail', read_timeout=120)
                transceivers = self.parse_nxos_transceiver(output, devicename)
            else:
                # IOS command
                output = conn.send_command('show interfaces transceiver', read_timeout=120)
                transceivers = self.parse_ios_transceiver(output, devicename)
            
            conn.disconnect()
            
            self.logger.info(f"Found {len(transceivers)} transceivers on {devicename}")
            
            # Thread-safe append
            with self.results_lock:
                self.results.extend(transceivers)
            
            return {'device': devicename, 'status': 'Success', 'count': len(transceivers)}
            
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"Failed to connect to {devicename} ({ip_addr}): {error_msg}")
            
            with self.results_lock:
                self.device_errors.append({
                    'devicename': devicename,
                    'ip_addr': ip_addr,
                    'error': error_msg
                })
            
            return {'device': devicename, 'status': 'Failed', 'error': error_msg}

    def collect_all_transceivers(self):
        """Collect transceiver information from all devices using threading"""
        self.logger.info(f"Starting transceiver collection from {len(self.devices)} devices using {self.MAX_WORKERS} threads")
        
        results_summary = []
        
        with ThreadPoolExecutor(max_workers=self.MAX_WORKERS) as executor:
            futures = {executor.submit(self.connect_and_collect, device): device for device in self.devices}
            
            for future in as_completed(futures):
                device = futures[future]
                try:
                    result = future.result()
                    results_summary.append(result)
                except Exception as e:
                    self.logger.error(f"Thread exception for {device.get('devicename', 'Unknown')}: {e}")
                    results_summary.append({
                        'device': device.get('devicename', 'Unknown'),
                        'status': 'Exception',
                        'error': str(e)
                    })
        
        self.logger.info(f"Collection complete. Found {len(self.results)} total transceivers")
        return results_summary

    def generate_excel_report(self, filename=None):
        """Generate Excel report with transceiver details and summary page"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.script_name}_report_{timestamp}.xlsx"
        
        self.logger.info(f"Generating Excel report: {filename}")
        
        wb = Workbook()
        
        # Create Summary sheet first
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary)
        
        # Create Transceiver Details sheet
        ws_details = wb.create_sheet("Transceiver Details")
        self._create_details_sheet(ws_details)
        
        # Create Failed Devices sheet if there are errors
        if self.device_errors:
            ws_errors = wb.create_sheet("Failed Devices")
            self._create_errors_sheet(ws_errors)
        
        # Save workbook
        wb.save(filename)
        
        # Set file permissions
        save_file_and_set_permissions(filename, show_info=True)
        
        self.logger.info(f"Excel report generated: {filename}")
        return filename

    def _create_summary_sheet(self, ws):
        """Create summary sheet with SFP counts by part number"""
        # Title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "Transceiver Inventory Summary"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = self.center_alignment
        
        # General stats
        ws['A3'] = "Total Devices Processed:"
        ws['B3'] = len(self.devices)
        ws['A4'] = "Successful Connections:"
        ws['B4'] = len(self.devices) - len(self.device_errors)
        ws['A5'] = "Failed Connections:"
        ws['B5'] = len(self.device_errors)
        ws['A6'] = "Total Transceivers Found:"
        ws['B6'] = len(self.results)
        
        # Style the stats
        for row in range(3, 7):
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).alignment = self.center_alignment
        
        # Part number summary header
        ws['A8'] = "SFP Count by Cisco Part Number"
        ws['A8'].font = Font(bold=True, size=12)
        
        # Headers for part number table
        headers = ['Cisco Part Number', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.fill = self.summary_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
        
        # Count by part number
        part_number_counts = Counter(t.get('cisco_part_number', 'Unknown') for t in self.results if t.get('cisco_part_number'))
        
        # Sort by count descending
        sorted_counts = sorted(part_number_counts.items(), key=lambda x: x[1], reverse=True)
        
        row = 10
        for part_number, count in sorted_counts:
            ws.cell(row=row, column=1, value=part_number).alignment = self.left_alignment
            ws.cell(row=row, column=2, value=count).alignment = self.center_alignment
            row += 1
        
        # Add total row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(self.results)).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = self.center_alignment
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_details_sheet(self, ws):
        """Create transceiver details sheet"""
        headers = [
            'Device Name',
            'Interface',
            'Cisco Part Number',
            'Serial Number',
            'Type',
            'Cisco Product ID',
            'Name'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add data rows - sort by device name then interface
        sorted_results = sorted(self.results, key=lambda x: (x.get('devicename', ''), x.get('interface', '')))
        
        for row_idx, record in enumerate(sorted_results, 2):
            ws.cell(row=row_idx, column=1, value=record.get('devicename', '')).alignment = self.left_alignment
            ws.cell(row=row_idx, column=2, value=record.get('interface', '')).alignment = self.left_alignment
            ws.cell(row=row_idx, column=3, value=record.get('cisco_part_number', '')).alignment = self.center_alignment
            ws.cell(row=row_idx, column=4, value=record.get('serial_number', '')).alignment = self.center_alignment
            ws.cell(row=row_idx, column=5, value=record.get('type', '')).alignment = self.left_alignment
            ws.cell(row=row_idx, column=6, value=record.get('cisco_product_id', '')).alignment = self.center_alignment
            ws.cell(row=row_idx, column=7, value=record.get('name', '')).alignment = self.left_alignment
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_errors_sheet(self, ws):
        """Create failed devices sheet"""
        headers = ['Device Name', 'IP Address', 'Error']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = self.header_font
            cell.alignment = self.center_alignment
        
        ws.freeze_panes = 'A2'
        
        for row_idx, error in enumerate(self.device_errors, 2):
            ws.cell(row=row_idx, column=1, value=error.get('devicename', '')).alignment = self.left_alignment
            ws.cell(row=row_idx, column=2, value=error.get('ip_addr', '')).alignment = self.center_alignment
            ws.cell(row=row_idx, column=3, value=error.get('error', '')).alignment = self.left_alignment
        
        self._auto_adjust_columns(ws)

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)  # Cap at 60 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def print_summary(self):
        """Print summary statistics to console"""
        total_devices = len(self.devices)
        successful = total_devices - len(self.device_errors)
        
        print(f"\n{'='*50}")
        print(f"  Transceiver Inventory Summary")
        print(f"{'='*50}")
        print(f"  Devices processed:      {total_devices}")
        print(f"  Successful connections: {successful}")
        print(f"  Failed connections:     {len(self.device_errors)}")
        print(f"  Total transceivers:     {len(self.results)}")
        
        if self.results:
            print(f"\n  SFP Count by Cisco Part Number:")
            print(f"  {'-'*40}")
            part_counts = Counter(t.get('cisco_part_number', 'Unknown') for t in self.results if t.get('cisco_part_number'))
            for part, count in sorted(part_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
                print(f"    {part:<30} {count:>5}")
            if len(part_counts) > 10:
                print(f"    ... and {len(part_counts) - 10} more part numbers")
        
        print(f"{'='*50}\n")


def main():
    """Main function"""
    # Check for command line argument for CSV file
    csv_file = "freedevices.csv"
    if len(sys.argv) > 1:
        csv_file = sys.argv[1]
    
    inventory = TransceiverInventory(csv_file=csv_file)
    
    # Load devices
    if not inventory.load_devices():
        print(f"Error: Could not load devices from {csv_file}")
        print("Expected CSV format: devicename,ipaddr,description,devicetype")
        print("devicetype should be: cisco_ios or cisco_nxos")
        return 1
    
    # Collect transceiver information
    inventory.collect_all_transceivers()
    
    # Generate Excel report
    report_filename = inventory.generate_excel_report()
    
    # Print summary
    inventory.print_summary()
    
    print(f"Report saved as: {report_filename}")
    
    return 0


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
