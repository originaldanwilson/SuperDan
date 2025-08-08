#!/usr/bin/env python3
"""
DNS Lookup Script for Windows - Hybrid Version with dnspython
Uses dnspython with Windows-optimized settings and multiple fallback methods
Specifically tuned for better internal IP resolution
Handles both internal (RFC 1918) and external IP addresses
Outputs results to Excel with proper formatting
"""

import csv
import socket
import ipaddress
import sys
import subprocess
import re
import platform
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import dns.resolver
import dns.reversename

def get_windows_dns_servers():
    """Get actual DNS servers from Windows using ipconfig"""
    dns_servers = []
    
    try:
        # Try ipconfig /all
        result = subprocess.run(['ipconfig', '/all'], capture_output=True, text=True, timeout=15, shell=True)
        if result.returncode == 0:
            lines = result.stdout.split('\n')
            
            for line in lines:
                line = line.strip()
                
                # Look for DNS Servers lines
                if ('DNS Servers' in line or 'DNS-Server' in line) and ':' in line:
                    # Extract IP after the colon
                    parts = line.split(':')
                    if len(parts) > 1:
                        ip = parts[1].strip()
                        # Validate it's an IPv4 address
                        try:
                            ipaddress.ip_address(ip)
                            if '.' in ip and ip != '127.0.0.1':  # IPv4 only, skip localhost
                                dns_servers.append(ip)
                        except ValueError:
                            pass
                # Sometimes DNS servers are on following lines (indented)
                elif line and len(line.split()) == 1 and not line.startswith(('   ', 'Ethernet', 'Wireless')):
                    try:
                        # Check if the line is just an IP address
                        ip = line.strip()
                        ipaddress.ip_address(ip)
                        if '.' in ip and ip != '127.0.0.1':  # IPv4, skip localhost
                            dns_servers.append(ip)
                    except ValueError:
                        pass
        
        if dns_servers:
            # Remove duplicates while preserving order
            seen = set()
            unique_servers = []
            for server in dns_servers:
                if server not in seen:
                    seen.add(server)
                    unique_servers.append(server)
            print(f"Found Windows DNS servers: {unique_servers}")
            return unique_servers
    
    except Exception as e:
        print(f"Could not get Windows DNS servers: {e}")
    
    # Fallback DNS servers
    print("Using fallback DNS servers")
    return ['8.8.8.8', '8.8.4.4', '1.1.1.1']

def is_rfc1918(ip_str):
    """Check if IP address is RFC 1918 (private/internal)"""
    try:
        ip = ipaddress.ip_address(ip_str)
        return ip.is_private
    except ValueError:
        return False

def dns_lookup_dnspython_aggressive(ip_address, dns_servers, is_internal=False):
    """
    Aggressive dnspython lookup with Windows and internal IP optimizations
    """
    attempts = []
    
    # For internal IPs, try more aggressive settings
    if is_internal:
        timeout_settings = [
            (20, 60),  # Very long timeout for internal
            (30, 90),  # Even longer
            (10, 30),  # Standard
        ]
    else:
        timeout_settings = [
            (15, 45),  # Standard external
            (25, 75),  # Longer external
            (10, 30),  # Short external
        ]
    
    for timeout, lifetime in timeout_settings:
        for dns_server in dns_servers[:3]:  # Try top 3 DNS servers
            try:
                resolver = dns.resolver.Resolver()
                resolver.nameservers = [dns_server]
                resolver.timeout = timeout
                resolver.lifetime = lifetime
                
                # For internal IPs, try different query settings
                if is_internal:
                    resolver.retry_count = 3  # More retries for internal
                    resolver.rotate = True    # Rotate through servers
                
                # Create reverse DNS name
                reverse_name = dns.reversename.from_address(ip_address)
                
                # Perform the lookup
                answer = resolver.resolve(reverse_name, 'PTR')
                hostname = str(answer[0]).rstrip('.')
                
                return hostname
                
            except dns.resolver.NXDOMAIN:
                attempts.append(f"NXDOMAIN via {dns_server}")
                continue
            except dns.resolver.Timeout:
                attempts.append(f"Timeout via {dns_server} ({timeout}s)")
                continue
            except Exception as e:
                attempts.append(f"Error via {dns_server}: {str(e)}")
                continue
    
    # Return the best error message
    if attempts:
        return f"No PTR record found (tried: {'; '.join(attempts[:3])})"
    return "No PTR record found"

def dns_lookup_nslookup_aggressive(ip_address, dns_servers, is_internal=False):
    """
    Aggressive nslookup with longer timeouts for internal IPs
    """
    timeout = 30 if is_internal else 20
    servers_to_try = [None] + dns_servers[:3]  # None = system default
    
    for dns_server in servers_to_try:
        try:
            if dns_server:
                cmd = ['nslookup', ip_address, dns_server]
            else:
                cmd = ['nslookup', ip_address]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, shell=True)
            
            if result.returncode == 0:
                lines = result.stdout.strip().split('\n')
                for line in lines:
                    line = line.strip()
                    
                    # Windows nslookup patterns:
                    if 'Name:' in line and line.startswith('Name:'):
                        hostname = line.split('Name:')[1].strip().rstrip('.')
                        if hostname and hostname != ip_address and not hostname.startswith('***'):
                            return hostname
                    
                    # Alternative pattern: "name = hostname"
                    elif 'name =' in line:
                        hostname = line.split('name =')[1].strip().rstrip('.')
                        if hostname and hostname != ip_address:
                            return hostname
                    
                    # Some Windows versions show hostname directly
                    elif (line and 
                          not line.startswith(('Server:', 'Address:', 'Non-authoritative', '***', 'Default', 'Aliases:', 'Name:', 'Addresses:')) and
                          not line.endswith('.in-addr.arpa') and
                          not 'can\'t find' in line.lower() and
                          not 'dns request timed out' in line.lower() and
                          not line.startswith('>')):
                        # Check if it looks like a hostname
                        if ('.' in line and 
                            not line.replace('.', '').replace('-', '').replace('_', '').isdigit() and
                            len(line) > 3):
                            return line.rstrip('.')
            
        except subprocess.TimeoutExpired:
            continue
        except Exception:
            continue
    
    return "No PTR record found"

def dns_lookup_socket_method(ip_address):
    """
    Socket method with longer timeout for Windows
    """
    try:
        socket.setdefaulttimeout(25)  # Longer timeout
        hostname, _, _ = socket.gethostbyaddr(ip_address)
        return hostname
    except socket.herror:
        return "No PTR record found"
    except socket.timeout:
        return "Socket timeout"
    except socket.gaierror as e:
        return f"Socket DNS error: {str(e)}"
    except Exception as e:
        return f"Socket error: {str(e)}"

def dns_lookup_combined_aggressive(ip_address, system_dns_servers, use_external_fallback=False):
    """
    Ultra-aggressive DNS lookup combining all methods with internal IP focus
    """
    is_internal = is_rfc1918(ip_address)
    
    # Determine which DNS servers to use
    if use_external_fallback and not is_internal:
        # For external addresses that fail with system DNS, try external DNS
        dns_servers = ['8.8.8.8', '8.8.4.4', '1.1.1.1', '1.0.0.1']
    else:
        # Use system DNS servers for internal addresses or first attempt on external
        dns_servers = system_dns_servers
    
    methods_tried = []
    
    # Method 1: Socket method (uses Windows DNS resolution) - prioritize for internal
    if not use_external_fallback:
        socket_result = dns_lookup_socket_method(ip_address)
        methods_tried.append(f"socket: {socket_result}")
        if not socket_result.startswith(('No PTR', 'Socket')):
            return socket_result
    
    # Method 2: Aggressive dnspython with optimized settings
    dnspython_result = dns_lookup_dnspython_aggressive(ip_address, dns_servers, is_internal)
    methods_tried.append(f"dnspython: {dnspython_result}")
    if not dnspython_result.startswith(('No PTR', 'DNS error:', 'DNS timeout')):
        return dnspython_result
    
    # Method 3: Aggressive nslookup with longer timeouts
    nslookup_result = dns_lookup_nslookup_aggressive(ip_address, dns_servers, is_internal)
    methods_tried.append(f"nslookup: {nslookup_result}")
    if not nslookup_result.startswith(('No PTR', 'DNS error')):
        return nslookup_result
    
    # For internal IPs that failed, try one more time with ALL available DNS servers
    if is_internal and not use_external_fallback:
        # Try with additional internal DNS servers that might exist
        extended_dns = dns_servers + ['192.168.1.1', '10.0.0.1', '172.16.0.1']
        extended_result = dns_lookup_dnspython_aggressive(ip_address, extended_dns, True)
        if not extended_result.startswith(('No PTR', 'DNS error:', 'DNS timeout')):
            return extended_result
    
    # Return the best result we got
    if socket_result and not socket_result.startswith('Socket'):
        return socket_result
    elif not dnspython_result.startswith('DNS error:') and dnspython_result != "No PTR record found":
        return dnspython_result
    else:
        return f"Failed all methods - Internal IP: {is_internal}"

def process_ip_batch(ip_list, system_dns_servers):
    """Process a batch of IP addresses with DNS lookups"""
    results = []
    
    # Separate internal and external IPs for different processing
    internal_ips = [ip for ip in ip_list if ip.strip() and is_rfc1918(ip.strip())]
    external_ips = [ip for ip in ip_list if ip.strip() and not is_rfc1918(ip.strip())]
    
    print(f"  - Internal IPs: {len(internal_ips)}")
    print(f"  - External IPs: {len(external_ips)}")
    
    # Process internal IPs with high-performance threading for massive scale
    with ThreadPoolExecutor(max_workers=40) as executor:  # High performance for internal
        future_to_ip = {}
        
        # Submit internal IPs first
        for ip in internal_ips:
            ip = ip.strip()
            if not ip:
                continue
            future = executor.submit(dns_lookup_combined_aggressive, ip, system_dns_servers, False)
            future_to_ip[future] = ip
        
        # Collect internal results
        for future in as_completed(future_to_ip):
            ip = future_to_ip[future]
            try:
                hostname = future.result()
                results.append((ip, hostname))
                print(f"Processed {ip} (internal): {hostname}")
            except Exception as e:
                results.append((ip, f"Error: {str(e)}"))
                print(f"Error processing {ip}: {str(e)}")
    
    # Process external IPs with high-performance threading
    with ThreadPoolExecutor(max_workers=50) as executor:
        future_to_ip = {}
        
        # Submit external IPs
        for ip in external_ips:
            ip = ip.strip()
            if not ip:
                continue
            future = executor.submit(dns_lookup_combined_aggressive, ip, system_dns_servers, False)
            future_to_ip[future] = ip
        
        # Collect external results
        for future in as_completed(future_to_ip):
            ip = future_to_ip[future]
            try:
                hostname = future.result()
                
                # If external address failed with system DNS, try external DNS
                if hostname.startswith(('Failed all methods', 'DNS error:', 'DNS timeout', 'Socket')):
                    hostname = dns_lookup_combined_aggressive(ip, system_dns_servers, True)
                
                results.append((ip, hostname))
                print(f"Processed {ip} (external): {hostname}")
            except Exception as e:
                results.append((ip, f"Error: {str(e)}"))
                print(f"Error processing {ip}: {str(e)}")
    
    return results

def read_ips_from_csv(csv_file):
    """Read IP addresses from CSV file"""
    ips = []
    try:
        with open(csv_file, 'r', newline='', encoding='utf-8') as file:
            # Try to detect if first row is header
            sample = file.read(1024)
            file.seek(0)
            sniffer = csv.Sniffer()
            try:
                has_header = sniffer.has_header(sample)
            except:
                has_header = False
            
            reader = csv.reader(file)
            
            if has_header:
                next(reader)  # Skip header row
            
            for row in reader:
                if row:  # Skip empty rows
                    # Take the first column that looks like an IP
                    for cell in row:
                        cell = cell.strip()
                        if cell:
                            try:
                                ipaddress.ip_address(cell)
                                ips.append(cell)
                                break
                            except ValueError:
                                continue
        
        print(f"Read {len(ips)} IP addresses from {csv_file}")
        return ips
        
    except FileNotFoundError:
        print(f"Error: CSV file '{csv_file}' not found")
        return []
    except Exception as e:
        print(f"Error reading CSV file: {str(e)}")
        return []

def save_to_excel(results, output_file):
    """Save results to Excel file with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "DNS Lookup Results"
    
    # Headers
    headers = ["IP Address", "Hostname"]
    ws.append(headers)
    
    # Format headers
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
    
    # Add data
    for ip, hostname in results:
        ws.append([ip, hostname])
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save file
    wb.save(output_file)
    print(f"Results saved to {output_file}")

def main():
    if len(sys.argv) != 3:
        print("Usage: python dns_lookup_windows_hybrid.py <input_csv_file> <output_excel_file>")
        print("Example: python dns_lookup_windows_hybrid.py ip_addresses.csv dns_results.xlsx")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_excel = sys.argv[2]
    
    system_name = platform.system()
    print(f"Starting AGGRESSIVE DNS lookup process on {system_name}...")
    print("=" * 80)
    print("This version uses:")
    print("- dnspython with ultra-aggressive settings for internal IPs")
    print("- Python socket library (Windows DNS)")
    print("- Windows nslookup with extended timeouts")
    print("- Separate processing for internal vs external IPs")
    print("- Extended DNS server lists for internal resolution")
    print("=" * 80)
    
    # Get actual system DNS servers
    system_dns_servers = get_windows_dns_servers()
    
    # Read IP addresses from CSV
    ips = read_ips_from_csv(input_csv)
    if not ips:
        print("No valid IP addresses found in the CSV file")
        sys.exit(1)
    
    # Count internal vs external
    internal_count = sum(1 for ip in ips if is_rfc1918(ip))
    external_count = len(ips) - internal_count
    
    print(f"Processing {len(ips)} IP addresses...")
    print(f"- Internal (RFC 1918) IPs: {internal_count}")
    print(f"- External IPs: {external_count}")
    print(f"Using system DNS servers: {system_dns_servers}")
    print("Internal IPs get ultra-aggressive treatment with extended timeouts!")
    print("=" * 80)
    
    start_time = time.time()
    
    # Process IPs in smaller batches for better control
    batch_size = 100  # Smaller batches for better monitoring
    all_results = []
    
    for i in range(0, len(ips), batch_size):
        batch = ips[i:i + batch_size]
        print(f"Processing batch {i//batch_size + 1} ({len(batch)} addresses)...")
        
        batch_results = process_ip_batch(batch, system_dns_servers)
        all_results.extend(batch_results)
        
        print(f"Completed batch {i//batch_size + 1}")
        print("-" * 60)
    
    elapsed_time = time.time() - start_time
    print(f"DNS lookups completed in {elapsed_time:.2f} seconds")
    print(f"Average: {elapsed_time/len(ips):.3f} seconds per IP")
    
    # Save to Excel
    print("Saving results to Excel...")
    save_to_excel(all_results, output_excel)
    
    # Detailed Summary
    successful_lookups = sum(1 for _, hostname in all_results 
                           if not hostname.startswith(('DNS error:', 'No PTR record', 'DNS timeout', 'Error:', 'Socket', 'Failed all methods')))
    
    internal_results = [(ip, hostname) for ip, hostname in all_results if is_rfc1918(ip)]
    external_results = [(ip, hostname) for ip, hostname in all_results if not is_rfc1918(ip)]
    
    internal_success = sum(1 for _, hostname in internal_results 
                          if not hostname.startswith(('DNS error:', 'No PTR record', 'DNS timeout', 'Error:', 'Socket', 'Failed all methods')))
    external_success = sum(1 for _, hostname in external_results 
                          if not hostname.startswith(('DNS error:', 'No PTR record', 'DNS timeout', 'Error:', 'Socket', 'Failed all methods')))
    
    print("=" * 80)
    print("DETAILED SUMMARY:")
    print(f"Total IP addresses processed: {len(all_results)}")
    print(f"Overall successful lookups: {successful_lookups}")
    print(f"Overall success rate: {(successful_lookups/len(all_results)*100):.1f}%")
    print("-" * 40)
    print(f"Internal IPs processed: {len(internal_results)}")
    print(f"Internal successful lookups: {internal_success}")
    print(f"Internal success rate: {(internal_success/len(internal_results)*100):.1f}%" if internal_results else "No internal IPs")
    print("-" * 40)
    print(f"External IPs processed: {len(external_results)}")
    print(f"External successful lookups: {external_success}")
    print(f"External success rate: {(external_success/len(external_results)*100):.1f}%" if external_results else "No external IPs")
    print("=" * 80)
    print(f"Results saved to: {output_excel}")
    print("DNS servers used:", system_dns_servers)

if __name__ == "__main__":
    main()
