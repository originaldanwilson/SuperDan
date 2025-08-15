#!/usr/bin/env python3
"""
DNS Lookup Script for IP Addresses - Built-in Libraries Only
Handles both internal (RFC 1918) and external IP addresses
Uses only Python standard library (no dnspython required)
Outputs results to Excel with proper formatting
"""

import csv
import socket
import ipaddress
import sys
import subprocess
import platform
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import time

def is_rfc1918(ip_str):
    """Check if IP address is RFC 1918 (private/internal)"""
    try:
        ip = ipaddress.ip_address(ip_str)
        return ip.is_private
    except ValueError:
        return False

def dns_lookup_socket(ip_address):
    """
    Perform reverse DNS lookup using Python's socket library
    
    Args:
        ip_address (str): IP address to lookup
    
    Returns:
        str: Hostname or error message
    """
    try:
        # Set socket timeout
        socket.setdefaulttimeout(10)
        
        # Perform reverse DNS lookup
        hostname, _, _ = socket.gethostbyaddr(ip_address)
        return hostname
        
    except socket.herror:
        return "No PTR record found"
    except socket.timeout:
        return "DNS timeout"
    except socket.gaierror as e:
        return f"DNS error: {str(e)}"
    except Exception as e:
        return f"Error: {str(e)}"

def dns_lookup_nslookup(ip_address, use_external_dns=False):
    """
    Perform reverse DNS lookup using nslookup command
    
    Args:
        ip_address (str): IP address to lookup
        use_external_dns (bool): Whether to use external DNS (8.8.8.8)
    
    Returns:
        str: Hostname or error message
    """
    try:
        # Construct nslookup command
        if use_external_dns:
            cmd = ['nslookup', ip_address, '8.8.8.8']
        else:
            cmd = ['nslookup', ip_address]
        
        # Execute nslookup with timeout
        result = subprocess.run(
            cmd, 
            capture_output=True, 
            text=True, 
            timeout=10
        )
        
        if result.returncode == 0:
            lines = result.stdout.strip().split('\n')
            for line in lines:
                line = line.strip()
                # Look for name = hostname pattern
                if 'name =' in line:
                    hostname = line.split('name =')[1].strip().rstrip('.')
                    return hostname
                # Alternative pattern for some nslookup outputs
                elif line.endswith('.in-addr.arpa'):
                    continue
                elif line and not line.startswith(('Server:', 'Address:', 'Non-authoritative')):
                    # Sometimes hostname is on its own line
                    if '.' in line and not line.startswith('**'):
                        return line.rstrip('.')
            
            return "No PTR record found"
        else:
            return "DNS lookup failed"
            
    except subprocess.TimeoutExpired:
        return "DNS timeout"
    except subprocess.CalledProcessError:
        return "DNS lookup failed"
    except FileNotFoundError:
        return "nslookup command not found"
    except Exception as e:
        return f"Error: {str(e)}"

def dns_lookup_combined(ip_address, use_external_dns=False):
    """
    Combined DNS lookup - try socket first, fallback to nslookup
    
    Args:
        ip_address (str): IP address to lookup
        use_external_dns (bool): Whether to use external DNS for nslookup fallback
    
    Returns:
        str: Hostname or error message
    """
    # First try Python's built-in socket method
    result = dns_lookup_socket(ip_address)
    
    # If socket method fails and we want external DNS, try nslookup
    if result.startswith(('DNS error:', 'No PTR record', 'DNS timeout', 'Error:')) and use_external_dns:
        nslookup_result = dns_lookup_nslookup(ip_address, use_external_dns=True)
        if not nslookup_result.startswith(('DNS', 'No PTR', 'Error:', 'nslookup')):
            return nslookup_result
    
    return result

def process_ip_batch(ip_list):
    """Process a batch of IP addresses with DNS lookups"""
    results = []
    
    with ThreadPoolExecutor(max_workers=50) as executor:
        # Submit all tasks
        future_to_ip = {}
        for ip in ip_list:
            ip = ip.strip()
            if not ip:
                continue
                
            # Determine if we should use external DNS for public IPs
            use_external_dns = not is_rfc1918(ip)
            future = executor.submit(dns_lookup_combined, ip, use_external_dns)
            future_to_ip[future] = ip
        
        # Collect results as they complete
        for future in as_completed(future_to_ip):
            ip = future_to_ip[future]
            try:
                hostname = future.result()
                results.append((ip, hostname))
                dns_type = "internal" if is_rfc1918(ip) else "external"
                print(f"Processed {ip} ({dns_type}): {hostname}")
            except Exception as e:
                results.append((ip, f"Error: {str(e)}"))
                print(f"Error processing {ip}: {str(e)}")
    
    return results

def read_ips_from_csv(csv_file):
    """Read IP addresses from CSV file"""
    ips = []
    try:
        with open(csv_file, 'r', newline='', encoding='utf-8') as file:
            # Try to detect if first row is header
            sample = file.read(1024)
            file.seek(0)
            sniffer = csv.Sniffer()
            has_header = sniffer.has_header(sample)
            
            reader = csv.reader(file)
            
            if has_header:
                next(reader)  # Skip header row
            
            for row in reader:
                if row:  # Skip empty rows
                    # Take the first column that looks like an IP
                    for cell in row:
                        cell = cell.strip()
                        if cell:
                            try:
                                ipaddress.ip_address(cell)
                                ips.append(cell)
                                break
                            except ValueError:
                                continue
        
        print(f"Read {len(ips)} IP addresses from {csv_file}")
        return ips
        
    except FileNotFoundError:
        print(f"Error: CSV file '{csv_file}' not found")
        return []
    except Exception as e:
        print(f"Error reading CSV file: {str(e)}")
        return []

def save_to_excel(results, output_file):
    """Save results to Excel file with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "DNS Lookup Results"
    
    # Headers
    headers = ["IP Address", "Hostname"]
    ws.append(headers)
    
    # Format headers
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
    
    # Add data
    for ip, hostname in results:
        ws.append([ip, hostname])
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save file
    wb.save(output_file)
    print(f"Results saved to {output_file}")

def main():
    if len(sys.argv) != 3:
        print("Usage: python3 dns_lookup_builtin.py <input_csv_file> <output_excel_file>")
        print("Example: python3 dns_lookup_builtin.py ip_addresses.csv dns_results.xlsx")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_excel = sys.argv[2]
    
    print("Starting DNS lookup process (using built-in libraries)...")
    print("=" * 50)
    
    # Check if nslookup is available for external DNS queries
    try:
        subprocess.run(['nslookup', '--version'], capture_output=True, timeout=5)
        print("nslookup available - will use for external DNS queries")
    except:
        print("nslookup not available - using socket library only")
    
    # Read IP addresses from CSV
    ips = read_ips_from_csv(input_csv)
    if not ips:
        print("No valid IP addresses found in the CSV file")
        sys.exit(1)
    
    print(f"Processing {len(ips)} IP addresses...")
    print("RFC 1918 (private) addresses will use system DNS")
    print("Public addresses will attempt external DNS lookups")
    print("=" * 50)
    
    start_time = time.time()
    
    # Process IPs in batches to manage memory and connections
    batch_size = 1000
    all_results = []
    
    for i in range(0, len(ips), batch_size):
        batch = ips[i:i + batch_size]
        print(f"Processing batch {i//batch_size + 1} ({len(batch)} addresses)...")
        
        batch_results = process_ip_batch(batch)
        all_results.extend(batch_results)
        
        print(f"Completed batch {i//batch_size + 1}")
        print("-" * 30)
    
    elapsed_time = time.time() - start_time
    print(f"DNS lookups completed in {elapsed_time:.2f} seconds")
    print(f"Average: {elapsed_time/len(ips):.3f} seconds per IP")
    
    # Save to Excel
    print("Saving results to Excel...")
    save_to_excel(all_results, output_excel)
    
    # Summary
    successful_lookups = sum(1 for _, hostname in all_results 
                           if not hostname.startswith(('DNS error:', 'No PTR record', 'DNS timeout', 'Error:')))
    
    print("=" * 50)
    print("SUMMARY:")
    print(f"Total IP addresses processed: {len(all_results)}")
    print(f"Successful DNS lookups: {successful_lookups}")
    print(f"Failed lookups: {len(all_results) - successful_lookups}")
    print(f"Success rate: {(successful_lookups/len(all_results)*100):.1f}%")
    print(f"Results saved to: {output_excel}")

if __name__ == "__main__":
    main()
