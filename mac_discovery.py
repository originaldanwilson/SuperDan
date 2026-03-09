#!/usr/bin/env python3
"""
MAC Address Discovery and DNS Resolution Tool

Connects to a Cisco NX-OS Layer 2 switch, collects locally learned MAC addresses
and VLAN numbers, resolves them to IP addresses via ARP on the Layer 3 switch,
performs DNS lookups, and generates an Excel report.

Usage:  python mac_discovery.py <l2_switch> <l3_switch>
Example: python mac_discovery.py dc-leaf-01 dc-spine-01

Sheet 1 (MAC-ARP-DNS) – full correlated data
Sheet 2 (IP Addresses) – IPs only, easy to copy en masse

Credentials come from tools.py.
"""

import argparse
import logging
import re
import subprocess
import sys
from datetime import datetime
from netmiko import ConnectHandler
from tools import get_netmiko_creds, getScriptName, setupLogging, save_file_and_set_permissions

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not found.  pip install openpyxl")
    sys.exit(1)

# ── Initialise ──────────────────────────────────────────────
scriptName = getScriptName()
setupLogging()
logger = logging.getLogger(__name__)
netmikoUser, passwd, enable = get_netmiko_creds()


# ── Device helpers ──────────────────────────────────────────
def connect_nxos(hostname):
    """Connect to an NX-OS device via netmiko."""
    logger.info(f"Connecting to {hostname}")
    try:
        conn = ConnectHandler(
            device_type="cisco_nxos",
            host=hostname,
            username=netmikoUser,
            password=passwd,
            secret=enable,
        )
        logger.info(f"Connected to {hostname}")
        return conn
    except Exception as e:
        logger.error(f"Failed to connect to {hostname}: {e}")
        print(f"  *** FAILED to connect to {hostname}: {e}")
        return None


def get_local_mac_addresses(conn, hostname):
    """Collect locally learned dynamic MAC addresses from an NX-OS L2 switch.

    Returns:
        list[dict]: [{mac, vlan, port}, ...]
    """
    logger.info(f"Running 'show mac address-table local' on {hostname}")
    output = conn.send_command("show mac address-table local")
    logger.debug(f"Output:\n{output}")

    entries = []
    # NX-OS format (typical):
    #  [*+~G] VLAN  MAC_Address          Type   age  Secure NTFY  Ports
    #  *      10    aabb.ccdd.0001       dynamic 0   F      F     Eth1/1
    pattern = re.compile(
        r'[\*\+~G]?\s*(\d+)\s+'                                        # VLAN
        r'([0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4})\s+'        # MAC
        r'(\w+)\s+'                                                     # Type
        r'\S+\s+'                                                       # age
        r'\S+\s+'                                                       # Secure
        r'\S+\s+'                                                       # NTFY
        r'(\S+)'                                                        # Port
    )

    for line in output.splitlines():
        m = pattern.search(line)
        if m and m.group(3).lower() == "dynamic":
            entries.append({
                "mac":  m.group(2).lower(),
                "vlan": m.group(1),
                "port": m.group(4),
            })

    logger.info(f"{len(entries)} dynamic local MACs on {hostname}")
    print(f"  {len(entries)} dynamic local MAC addresses found")
    return entries


def get_arp_table(conn, hostname):
    """Collect ARP table from an NX-OS L3 switch.

    Returns:
        dict: {mac_address: {ip, interface}, ...}
    """
    logger.info(f"Running 'show ip arp' on {hostname}")
    output = conn.send_command("show ip arp")
    logger.debug(f"Output:\n{output}")

    arp = {}
    # NX-OS format:
    #  Address       Age       MAC Address          Interface    Flags
    #  10.1.1.1      00:05:32  aabb.ccdd.0001       Vlan10
    pattern = re.compile(
        r'(\d+\.\d+\.\d+\.\d+)\s+'                                     # IP
        r'\S+\s+'                                                       # Age
        r'([0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4})\s+'        # MAC
        r'(\S+)'                                                        # Interface
    )

    for line in output.splitlines():
        m = pattern.search(line)
        if m:
            arp[m.group(2).lower()] = {"ip": m.group(1), "interface": m.group(3)}

    logger.info(f"{len(arp)} ARP entries on {hostname}")
    print(f"  {len(arp)} ARP entries found")
    return arp


# ── DNS lookup ──────────────────────────────────────────────
def do_nslookup(ip):
    """Run nslookup and return the resolved hostname, or 'N/A'."""
    try:
        result = subprocess.run(
            ["nslookup", ip],
            capture_output=True, text=True, timeout=10,
        )
        m = re.search(r"Name:\s+(\S+)", result.stdout)
        if m:
            return m.group(1)
    except (subprocess.TimeoutExpired, Exception) as e:
        logger.warning(f"nslookup failed for {ip}: {e}")
    return "N/A"


# ── Correlate ───────────────────────────────────────────────
def correlate(mac_entries, arp, l2_switch, l3_switch):
    """Match MAC -> IP (ARP) -> DNS and return combined rows."""
    rows = []
    for entry in mac_entries:
        mac  = entry["mac"]
        vlan = entry["vlan"]
        port = entry["port"]

        arp_hit  = arp.get(mac, {})
        ip       = arp_hit.get("ip", "N/A")
        arp_intf = arp_hit.get("interface", "N/A")

        dns = "N/A"
        if ip != "N/A":
            print(f"    nslookup {ip} ... ", end="", flush=True)
            dns = do_nslookup(ip)
            print(dns)

        rows.append({
            "mac": mac, "vlan": vlan, "port": port,
            "ip": ip, "arp_interface": arp_intf, "dns": dns,
            "l2_switch": l2_switch, "l3_switch": l3_switch,
        })
    return rows


# ── Excel output ────────────────────────────────────────────
def write_excel(rows, l2_switch):
    """Write results to Excel with two sheets.

    Sheet 1 – MAC-ARP-DNS  (full data)
    Sheet 2 – IP Addresses (just IPs for easy copy)
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"mac_discovery_{l2_switch}_{ts}.xlsx"
    wb = Workbook()

    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    center   = Alignment(horizontal="center", vertical="center")

    # ── Sheet 1: MAC-ARP-DNS ──
    ws = wb.active
    ws.title = "MAC-ARP-DNS"
    headers = [
        "MAC Address", "VLAN", "Port", "IP Address",
        "ARP Interface", "DNS Name", "L2 Switch", "L3 Switch",
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center

    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=row["mac"])
        ws.cell(row=r, column=2, value=row["vlan"])
        ws.cell(row=r, column=3, value=row["port"])
        ws.cell(row=r, column=4, value=row["ip"])
        ws.cell(row=r, column=5, value=row["arp_interface"])
        ws.cell(row=r, column=6, value=row["dns"])
        ws.cell(row=r, column=7, value=row["l2_switch"])
        ws.cell(row=r, column=8, value=row["l3_switch"])

    # Auto-width columns
    for c in range(1, len(headers) + 1):
        max_w = len(headers[c - 1])
        for r in range(2, len(rows) + 2):
            v = ws.cell(row=r, column=c).value
            if v and len(str(v)) > max_w:
                max_w = len(str(v))
        ws.column_dimensions[get_column_letter(c)].width = max_w + 3
    ws.freeze_panes = "A2"

    # ── Sheet 2: IP Addresses (easy copy) ──
    ws2 = wb.create_sheet(title="IP Addresses")
    cell = ws2.cell(row=1, column=1, value="IP Address")
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = center

    r = 2
    for row in rows:
        if row["ip"] != "N/A":
            ws2.cell(row=r, column=1, value=row["ip"])
            r += 1
    ws2.column_dimensions["A"].width = 20
    ws2.freeze_panes = "A2"

    wb.save(filename)
    save_file_and_set_permissions(filename)
    logger.info(f"Report saved: {filename}")
    print(f"\n  Report saved: {filename}")
    return filename


# ── Main ────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Discover local MACs on an NX-OS L2 switch, resolve to IP/DNS via L3 ARP, export to Excel."
    )
    parser.add_argument("l2_switch", help="Hostname/IP of the Layer 2 NX-OS switch")
    parser.add_argument("l3_switch", help="Hostname/IP of the Layer 3 NX-OS switch")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"MAC Discovery & DNS Resolution")
    print(f"  L2 Switch: {args.l2_switch}")
    print(f"  L3 Switch: {args.l3_switch}")
    print(f"{'='*60}")

    # 1 – L2: collect local MAC addresses
    print(f"\n[1/4] Collecting MACs from {args.l2_switch}")
    l2 = connect_nxos(args.l2_switch)
    if not l2:
        sys.exit(1)
    try:
        macs = get_local_mac_addresses(l2, args.l2_switch)
    finally:
        l2.disconnect()
        logger.info(f"Disconnected from {args.l2_switch}")

    if not macs:
        print("  No local dynamic MACs found. Nothing to do.")
        sys.exit(0)

    # 2 – L3: collect ARP table
    print(f"\n[2/4] Collecting ARP from {args.l3_switch}")
    l3 = connect_nxos(args.l3_switch)
    if not l3:
        sys.exit(1)
    try:
        arp = get_arp_table(l3, args.l3_switch)
    finally:
        l3.disconnect()
        logger.info(f"Disconnected from {args.l3_switch}")

    # 3 – Correlate MAC→IP  and run DNS lookups
    print(f"\n[3/4] DNS lookups")
    rows = correlate(macs, arp, args.l2_switch, args.l3_switch)

    resolved_ip  = sum(1 for r in rows if r["ip"]  != "N/A")
    resolved_dns = sum(1 for r in rows if r["dns"] != "N/A")

    # 4 – Write Excel
    print(f"\n[4/4] Writing report")
    filename = write_excel(rows, args.l2_switch)

    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"  MACs found:      {len(macs)}")
    print(f"  ARP resolved:    {resolved_ip}")
    print(f"  DNS resolved:    {resolved_dns}")
    print(f"  Report:          {filename}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
