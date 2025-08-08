#!/usr/bin/env python3
"""
DNS Lookup Script for IP Addresses - Using dig +short -x
Uses dig command for reliable reverse DNS lookups
Handles both internal (RFC 1918) and external IP addresses
Outputs results to Excel with proper formatting
"""

import csv
import socket
import ipaddress
import sys
import subprocess
import re
import platform
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import dns.resolver
import dns.reversename
import time

def get_system_dns_servers():
    """Get DNS servers based on the operating system"""
    system = platform.system().lower()
    dns_servers = []
    
    if system == 'windows':
        try:
            # Try ipconfig /all for Windows
            result = subprocess.run(['ipconfig', '/all'], capture_output=True, text=True, timeout=15, shell=True)
            if result.returncode == 0:
                lines = result.stdout.split('\n')
                for line in lines:
                    line = line.strip()
                    if 'DNS Servers' in line or 'DNS-Server' in line:
                        parts = line.split(':')
                        if len(parts) > 1:
                            ip = parts[1].strip()
                            try:
                                ipaddress.ip_address(ip)
                                if '.' in ip:  # IPv4 only
                                    dns_servers.append(ip)
                            except ValueError:
                                pass
        except Exception:
            pass
    
    elif system == 'linux':
        try:
            # Try resolvectl for systemd-resolved systems
            result = subprocess.run(['resolvectl', 'status'], capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                lines = result.stdout.split('\n')
                for line in lines:
                    if 'DNS Servers:' in line:
                        servers = line.split('DNS Servers:')[1].strip()
                        for server in servers.split():
                            if '.' in server and ':' not in server:
                                dns_servers.append(server)
                    elif 'Current DNS Server:' in line:
                        server = line.split('Current DNS Server:')[1].strip()
                        if '.' in server and ':' not in server:
                            dns_servers.insert(0, server)
        except Exception:
            pass
    
    if dns_servers:
        # Remove duplicates while preserving order
        seen = set()
        unique_servers = []
        for server in dns_servers:
            if server not in seen:
                seen.add(server)
                unique_servers.append(server)
        print(f"Found system DNS servers: {unique_servers}")
        return unique_servers
    
    # Fallback to common DNS servers
    print("Using fallback DNS servers")
    return ['8.8.8.8', '8.8.4.4']

def is_rfc1918(ip_str):
    """Check if IP address is RFC 1918 (private/internal)"""
    try:
        ip = ipaddress.ip_address(ip_str)
        return ip.is_private
    except ValueError:
        return False

def dns_lookup_dig(ip_address, dns_servers=None, retry_count=2):
    """
    Perform reverse DNS lookup using dig +short -x command
    This is often the most reliable method
    """
    for attempt in range(retry_count + 1):
        # Try with system DNS first, then with specific DNS servers
        dns_server_list = [None]  # None means use system default
        if dns_servers:
            dns_server_list.extend(dns_servers[:3])  # Try up to 3 DNS servers
        
        for dns_server in dns_server_list:
            try:
                if dns_server:
                    cmd = ['dig', '+short', '-x', ip_address, f'@{dns_server}']
                else:
                    cmd = ['dig', '+short', '-x', ip_address]
                
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                
                if result.returncode == 0 and result.stdout.strip():
                    hostname = result.stdout.strip()
                    # dig +short can return multiple lines, take the first valid one
                    lines = hostname.split('\n')
                    for line in lines:
                        line = line.strip().rstrip('.')
                        if line and not line.startswith((';;', ';')):
                            # Validate it looks like a hostname
                            if line and '.' in line and not line.replace('.', '').isdigit():
                                return line
                
            except subprocess.TimeoutExpired:
                continue
            except FileNotFoundError:
                # dig command not found, will fall back to other methods
                return None
            except Exception as e:
                continue
    
    return "No PTR record found"

def dns_lookup_nslookup_style(ip_address, dns_servers, retry_count=2):
    """
    Fallback: Perform DNS lookup using nslookup command
    """
    for attempt in range(retry_count + 1):
        for dns_server in dns_servers[:2]:  # Try first 2 DNS servers
            try:
                cmd = ['nslookup', ip_address, dns_server]
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                
                if result.returncode == 0:
                    lines = result.stdout.strip().split('\n')
                    for line in lines:
                        line = line.strip()
                        # Look for various nslookup output patterns
                        if 'name =' in line:
                            hostname = line.split('name =')[1].strip().rstrip('.')
                            return hostname
                        elif line.startswith('Name:'):
                            hostname = line.split('Name:')[1].strip().rstrip('.')
                            if hostname and hostname != ip_address:
                                return hostname
                
            except subprocess.TimeoutExpired:
                continue
            except Exception:
                continue
    
    return "No PTR record found"

def dns_lookup_dnspython(ip_address, dns_servers, retry_count=2):
    """
    Fallback: Perform reverse DNS lookup using dnspython
    """
    last_error = None
    
    for attempt in range(retry_count + 1):
        try:
            resolver = dns.resolver.Resolver()
            resolver.nameservers = dns_servers
            resolver.timeout = 15
            resolver.lifetime = 30
            
            reverse_name = dns.reversename.from_address(ip_address)
            answer = resolver.resolve(reverse_name, 'PTR')
            hostname = str(answer[0]).rstrip('.')
            
            return hostname
            
        except dns.resolver.NXDOMAIN:
            return "No PTR record found"
        except dns.resolver.Timeout as e:
            last_error = f"DNS timeout (attempt {attempt + 1})"
            if attempt < retry_count:
                time.sleep(1)
                continue
        except Exception as e:
            last_error = f"DNS error: {str(e)} (attempt {attempt + 1})"
            if attempt < retry_count:
                time.sleep(1)
                continue
    
    return last_error if last_error else "DNS lookup failed"

def dns_lookup_socket_method(ip_address):
    """
    Fallback: Perform DNS lookup using Python's socket library
    """
    try:
        socket.setdefaulttimeout(15)
        hostname, _, _ = socket.gethostbyaddr(ip_address)
        return hostname
    except socket.herror:
        return "No PTR record found"
    except socket.timeout:
        return "Socket timeout"
    except Exception as e:
        return f"Socket error: {str(e)}"

def dns_lookup_combined(ip_address, system_dns_servers, use_external_fallback=False):
    """
    Combined DNS lookup - try dig first, then fallback methods
    """
    # Determine which DNS servers to use
    if use_external_fallback and not is_rfc1918(ip_address):
        # For external addresses that fail with system DNS, try external DNS
        dns_servers = ['8.8.8.8', '8.8.4.4', '1.1.1.1', '1.0.0.1']
    else:
        # Use system DNS servers for internal addresses or first attempt on external
        dns_servers = system_dns_servers
    
    # Method 1: Try dig +short -x (most reliable)
    dig_result = dns_lookup_dig(ip_address, dns_servers)
    if dig_result and not dig_result.startswith(('No PTR', 'DNS error')):
        return dig_result
    elif dig_result is None:
        # dig command not available, skip to other methods
        pass
    
    # Method 2: Try socket method (uses system DNS naturally)
    if not use_external_fallback:  # Only for first attempt with system DNS
        socket_result = dns_lookup_socket_method(ip_address)
        if not socket_result.startswith(('No PTR', 'Socket')):
            return socket_result
    
    # Method 3: Try dnspython
    dnspython_result = dns_lookup_dnspython(ip_address, dns_servers)
    if not dnspython_result.startswith(('DNS error:', 'DNS timeout', 'No nameservers')):
        return dnspython_result
    
    # Method 4: Try nslookup as last resort
    nslookup_result = dns_lookup_nslookup_style(ip_address, dns_servers)
    if not nslookup_result.startswith(('No PTR', 'DNS error')):
        return nslookup_result
    
    # Return the best error message we got
    if dig_result and dig_result != "No PTR record found":
        return dig_result
    elif not dnspython_result.startswith('DNS error:'):
        return dnspython_result
    else:
        return "No PTR record found"

def process_ip_batch(ip_list, system_dns_servers):
    """Process a batch of IP addresses with DNS lookups"""
    results = []
    
    with ThreadPoolExecutor(max_workers=15) as executor:
        # Submit all tasks
        future_to_ip = {}
        for ip in ip_list:
            ip = ip.strip()
            if not ip:
                continue
                
            future = executor.submit(dns_lookup_combined, ip, system_dns_servers, False)
            future_to_ip[future] = ip
        
        # Collect results as they complete
        for future in as_completed(future_to_ip):
            ip = future_to_ip[future]
            try:
                hostname = future.result()
                
                # If external address failed with system DNS, try external DNS
                if (hostname.startswith(('DNS error:', 'DNS timeout', 'No nameservers', 'Socket')) and 
                    not is_rfc1918(ip)):
                    hostname = dns_lookup_combined(ip, system_dns_servers, True)
                
                results.append((ip, hostname))
                dns_type = "internal" if is_rfc1918(ip) else "external"
                print(f"Processed {ip} ({dns_type}): {hostname}")
            except Exception as e:
                results.append((ip, f"Error: {str(e)}"))
                print(f"Error processing {ip}: {str(e)}")
    
    return results

def read_ips_from_csv(csv_file):
    """Read IP addresses from CSV file"""
    ips = []
    try:
        with open(csv_file, 'r', newline='', encoding='utf-8') as file:
            # Try to detect if first row is header
            sample = file.read(1024)
            file.seek(0)
            sniffer = csv.Sniffer()
            has_header = sniffer.has_header(sample)
            
            reader = csv.reader(file)
            
            if has_header:
                next(reader)  # Skip header row
            
            for row in reader:
                if row:  # Skip empty rows
                    # Take the first column that looks like an IP
                    for cell in row:
                        cell = cell.strip()
                        if cell:
                            try:
                                ipaddress.ip_address(cell)
                                ips.append(cell)
                                break
                            except ValueError:
                                continue
        
        print(f"Read {len(ips)} IP addresses from {csv_file}")
        return ips
        
    except FileNotFoundError:
        print(f"Error: CSV file '{csv_file}' not found")
        return []
    except Exception as e:
        print(f"Error reading CSV file: {str(e)}")
        return []

def save_to_excel(results, output_file):
    """Save results to Excel file with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "DNS Lookup Results"
    
    # Headers
    headers = ["IP Address", "Hostname"]
    ws.append(headers)
    
    # Format headers
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
    
    # Add data
    for ip, hostname in results:
        ws.append([ip, hostname])
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save file
    wb.save(output_file)
    print(f"Results saved to {output_file}")

def check_dig_availability():
    """Check if dig command is available"""
    try:
        result = subprocess.run(['dig', '-v'], capture_output=True, text=True, timeout=5)
        return True
    except FileNotFoundError:
        return False
    except Exception:
        return False

def main():
    if len(sys.argv) != 3:
        print("Usage: python3 dns_lookup_dig.py <input_csv_file> <output_excel_file>")
        print("Example: python3 dns_lookup_dig.py ip_addresses.csv dns_results.xlsx")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_excel = sys.argv[2]
    
    system_name = platform.system()
    print(f"Starting DNS lookup process on {system_name} using dig +short -x...")
    print("=" * 70)
    
    # Check if dig is available
    dig_available = check_dig_availability()
    if dig_available:
        print("✓ dig command is available - using as primary method")
    else:
        print("⚠ dig command not found - will use fallback methods only")
    
    # Get actual system DNS servers
    system_dns_servers = get_system_dns_servers()
    
    # Read IP addresses from CSV
    ips = read_ips_from_csv(input_csv)
    if not ips:
        print("No valid IP addresses found in the CSV file")
        sys.exit(1)
    
    print(f"Processing {len(ips)} IP addresses...")
    print(f"Using system DNS servers: {system_dns_servers}")
    print("Primary method: dig +short -x")
    print("Fallback methods: socket, dnspython, nslookup")
    print("RFC 1918 (private) addresses will use system DNS")
    print("Public addresses will use system DNS first, external DNS as fallback")
    print("=" * 70)
    
    start_time = time.time()
    
    # Process IPs in batches
    batch_size = 500
    all_results = []
    
    for i in range(0, len(ips), batch_size):
        batch = ips[i:i + batch_size]
        print(f"Processing batch {i//batch_size + 1} ({len(batch)} addresses)...")
        
        batch_results = process_ip_batch(batch, system_dns_servers)
        all_results.extend(batch_results)
        
        print(f"Completed batch {i//batch_size + 1}")
        print("-" * 50)
    
    elapsed_time = time.time() - start_time
    print(f"DNS lookups completed in {elapsed_time:.2f} seconds")
    print(f"Average: {elapsed_time/len(ips):.3f} seconds per IP")
    
    # Save to Excel
    print("Saving results to Excel...")
    save_to_excel(all_results, output_excel)
    
    # Summary
    successful_lookups = sum(1 for _, hostname in all_results 
                           if not hostname.startswith(('DNS error:', 'No PTR record', 'DNS timeout', 'Error:', 'Socket')))
    
    print("=" * 70)
    print("SUMMARY:")
    print(f"Total IP addresses processed: {len(all_results)}")
    print(f"Successful DNS lookups: {successful_lookups}")
    print(f"Failed lookups: {len(all_results) - successful_lookups}")
    print(f"Success rate: {(successful_lookups/len(all_results)*100):.1f}%")
    print(f"Results saved to: {output_excel}")
    print("=" * 70)
    print("System:", system_name)
    print("DNS servers used:", system_dns_servers)
    print("Primary method: dig +short -x" + (" (available)" if dig_available else " (not available)"))

if __name__ == "__main__":
    main()
