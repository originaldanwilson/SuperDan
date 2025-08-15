#!/usr/bin/env python3
"""
DNS Lookup Script for IP Addresses - Windows Compatible
Uses actual Windows DNS servers and resolution methods
Handles both internal (RFC 1918) and external IP addresses
Outputs results to Excel with proper formatting
"""

import csv
import socket
import ipaddress
import sys
import subprocess
import re
import platform
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import dns.resolver
import dns.reversename
import time

def get_windows_dns_servers():
    """Get actual DNS servers from Windows using ipconfig or netsh"""
    dns_servers = []
    
    try:
        # Try ipconfig /all first (most common)
        result = subprocess.run(['ipconfig', '/all'], capture_output=True, text=True, timeout=15, shell=True)
        if result.returncode == 0:
            lines = result.stdout.split('\n')
            for line in lines:
                line = line.strip()
                # Look for DNS Servers lines
                if 'DNS Servers' in line or 'DNS-Server' in line:
                    # Extract IP after the colon
                    parts = line.split(':')
                    if len(parts) > 1:
                        ip = parts[1].strip()
                        # Validate it's an IPv4 address
                        try:
                            ipaddress.ip_address(ip)
                            if '.' in ip:  # IPv4 only
                                dns_servers.append(ip)
                        except ValueError:
                            pass
                # Sometimes DNS servers are on following lines with just IP
                elif line and not line.startswith(('   ', 'Ethernet', 'Wireless', 'Connection')):
                    try:
                        # Check if the line is just an IP address
                        parts = line.split()
                        for part in parts:
                            ipaddress.ip_address(part)
                            if '.' in part:  # IPv4
                                dns_servers.append(part)
                    except ValueError:
                        pass
        
        if dns_servers:
            # Remove duplicates while preserving order
            seen = set()
            unique_servers = []
            for server in dns_servers:
                if server not in seen:
                    seen.add(server)
                    unique_servers.append(server)
            print(f"Found Windows DNS servers: {unique_servers}")
            return unique_servers
    
    except Exception as e:
        print(f"Could not get Windows DNS servers with ipconfig: {e}")
    
    # Fallback: try netsh
    try:
        result = subprocess.run(['netsh', 'interface', 'ip', 'show', 'dns'], 
                              capture_output=True, text=True, timeout=15, shell=True)
        if result.returncode == 0:
            lines = result.stdout.split('\n')
            for line in lines:
                if 'Statically Configured DNS Servers:' in line or 'DNS servers configured through DHCP:' in line:
                    continue
                # Look for IP addresses
                line = line.strip()
                try:
                    ipaddress.ip_address(line)
                    if '.' in line:  # IPv4
                        dns_servers.append(line)
                except ValueError:
                    pass
        
        if dns_servers:
            print(f"Found Windows DNS servers via netsh: {dns_servers}")
            return dns_servers
    
    except Exception as e:
        print(f"Could not get Windows DNS servers with netsh: {e}")
    
    # Final fallback: use common Windows DNS or system defaults
    print("Using fallback DNS servers")
    return ['8.8.8.8', '8.8.4.4']

def get_system_dns_servers():
    """Get DNS servers based on the operating system"""
    system = platform.system().lower()
    
    if system == 'windows':
        return get_windows_dns_servers()
    elif system == 'linux':
        # Use the Linux method from before
        try:
            result = subprocess.run(['resolvectl', 'status'], capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                dns_servers = []
                lines = result.stdout.split('\n')
                for line in lines:
                    if 'DNS Servers:' in line:
                        servers = line.split('DNS Servers:')[1].strip()
                        for server in servers.split():
                            if '.' in server and ':' not in server:
                                dns_servers.append(server)
                    elif 'Current DNS Server:' in line:
                        server = line.split('Current DNS Server:')[1].strip()
                        if '.' in server and ':' not in server:
                            dns_servers.insert(0, server)
                
                if dns_servers:
                    print(f"Found Linux DNS servers: {dns_servers}")
                    return dns_servers
        except Exception:
            pass
    
    # Universal fallback
    return ['8.8.8.8', '8.8.4.4']

def is_rfc1918(ip_str):
    """Check if IP address is RFC 1918 (private/internal)"""
    try:
        ip = ipaddress.ip_address(ip_str)
        return ip.is_private
    except ValueError:
        return False

def dns_lookup_nslookup_style(ip_address, dns_servers, retry_count=2):
    """
    Perform DNS lookup using nslookup command directly to match manual behavior
    Works on both Windows and Linux
    """
    for attempt in range(retry_count + 1):
        for dns_server in dns_servers:
            try:
                cmd = ['nslookup', ip_address, dns_server]
                # On Windows, don't use shell=True unless necessary
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                
                if result.returncode == 0:
                    lines = result.stdout.strip().split('\n')
                    for line in lines:
                        line = line.strip()
                        # Look for name = hostname pattern (common format)
                        if 'name =' in line:
                            hostname = line.split('name =')[1].strip().rstrip('.')
                            return hostname
                        # Windows sometimes shows "Name:" format
                        elif line.startswith('Name:'):
                            hostname = line.split('Name:')[1].strip().rstrip('.')
                            if hostname and hostname != ip_address:
                                return hostname
                        # Alternative pattern: sometimes the hostname appears directly
                        elif (line and 
                              not line.startswith(('Server:', 'Address:', 'Non-authoritative', '**', 'Default', 'Name:', 'Aliases:')) and
                              not line.endswith('.in-addr.arpa') and
                              not line.startswith(('Can\'t', 'NXDOMAIN', 'connection timed out', '*** ', 'DNS request timed out'))):
                            # Check if it looks like a hostname
                            if '.' in line and not line.replace('.', '').replace('-', '').replace('_', '').isdigit():
                                return line.rstrip('.')
                
            except subprocess.TimeoutExpired:
                continue
            except Exception as e:
                print(f"nslookup error for {ip_address}: {e}")
                continue
    
    return "No PTR record found"

def dns_lookup_dnspython(ip_address, dns_servers, retry_count=2):
    """
    Perform reverse DNS lookup using dnspython with system DNS servers
    """
    last_error = None
    
    for attempt in range(retry_count + 1):
        try:
            # Create a new resolver for each lookup
            resolver = dns.resolver.Resolver()
            resolver.nameservers = dns_servers
            
            # Set longer timeouts for better reliability
            resolver.timeout = 15  # Individual query timeout
            resolver.lifetime = 30  # Total lifetime for all queries
            
            # Create reverse DNS name
            reverse_name = dns.reversename.from_address(ip_address)
            
            # Perform the lookup
            answer = resolver.resolve(reverse_name, 'PTR')
            hostname = str(answer[0]).rstrip('.')
            
            return hostname
            
        except dns.resolver.NXDOMAIN:
            return "No PTR record found"
        except dns.resolver.Timeout as e:
            last_error = f"DNS timeout (attempt {attempt + 1})"
            if attempt < retry_count:
                time.sleep(1)  # Brief pause before retry
                continue
        except dns.resolver.NoNameservers as e:
            last_error = f"No nameservers available (attempt {attempt + 1})"
            if attempt < retry_count:
                time.sleep(1)
                continue
        except Exception as e:
            last_error = f"DNS error: {str(e)} (attempt {attempt + 1})"
            if attempt < retry_count:
                time.sleep(1)
                continue
    
    return last_error if last_error else "DNS lookup failed"

def dns_lookup_socket_method(ip_address):
    """
    Perform DNS lookup using Python's socket library (most compatible)
    """
    try:
        socket.setdefaulttimeout(15)
        hostname, _, _ = socket.gethostbyaddr(ip_address)
        return hostname
    except socket.herror:
        return "No PTR record found"
    except socket.timeout:
        return "Socket timeout"
    except socket.gaierror as e:
        return f"Socket DNS error: {str(e)}"
    except Exception as e:
        return f"Socket error: {str(e)}"

def dns_lookup_combined(ip_address, system_dns_servers, use_external_fallback=False):
    """
    Combined DNS lookup - try multiple methods for best compatibility
    """
    # Determine which DNS servers to use
    if use_external_fallback and not is_rfc1918(ip_address):
        # For external addresses that fail with system DNS, try external DNS
        dns_servers = ['8.8.8.8', '8.8.4.4', '1.1.1.1', '1.0.0.1']
    else:
        # Use system DNS servers for internal addresses or first attempt on external
        dns_servers = system_dns_servers
    
    # Try socket method first (most compatible with system DNS)
    if not use_external_fallback:
        socket_result = dns_lookup_socket_method(ip_address)
        if not socket_result.startswith(('No PTR', 'Socket timeout', 'Socket DNS error', 'Socket error')):
            return socket_result
    
    # Try dnspython method
    dnspython_result = dns_lookup_dnspython(ip_address, dns_servers)
    if not dnspython_result.startswith(('DNS error:', 'DNS timeout', 'No nameservers')):
        return dnspython_result
    
    # Finally try nslookup for maximum compatibility
    nslookup_result = dns_lookup_nslookup_style(ip_address, dns_servers)
    if not nslookup_result.startswith(('No PTR', 'DNS error')):
        return nslookup_result
    
    # Return the best error message
    if socket_result and not socket_result.startswith('Socket'):
        return socket_result
    elif not dnspython_result.startswith('DNS error:'):
        return dnspython_result
    else:
        return nslookup_result

def process_ip_batch(ip_list, system_dns_servers):
    """Process a batch of IP addresses with DNS lookups"""
    results = []
    
    with ThreadPoolExecutor(max_workers=10) as executor:  # Very conservative for Windows
        # Submit all tasks
        future_to_ip = {}
        for ip in ip_list:
            ip = ip.strip()
            if not ip:
                continue
                
            future = executor.submit(dns_lookup_combined, ip, system_dns_servers, False)
            future_to_ip[future] = ip
        
        # Collect results as they complete
        for future in as_completed(future_to_ip):
            ip = future_to_ip[future]
            try:
                hostname = future.result()
                
                # If external address failed with system DNS, try external DNS
                if (hostname.startswith(('DNS error:', 'DNS timeout', 'No nameservers', 'Socket')) and 
                    not is_rfc1918(ip)):
                    hostname = dns_lookup_combined(ip, system_dns_servers, True)
                
                results.append((ip, hostname))
                dns_type = "internal" if is_rfc1918(ip) else "external"
                print(f"Processed {ip} ({dns_type}): {hostname}")
            except Exception as e:
                results.append((ip, f"Error: {str(e)}"))
                print(f"Error processing {ip}: {str(e)}")
    
    return results

def read_ips_from_csv(csv_file):
    """Read IP addresses from CSV file"""
    ips = []
    try:
        with open(csv_file, 'r', newline='', encoding='utf-8') as file:
            # Try to detect if first row is header
            sample = file.read(1024)
            file.seek(0)
            sniffer = csv.Sniffer()
            has_header = sniffer.has_header(sample)
            
            reader = csv.reader(file)
            
            if has_header:
                next(reader)  # Skip header row
            
            for row in reader:
                if row:  # Skip empty rows
                    # Take the first column that looks like an IP
                    for cell in row:
                        cell = cell.strip()
                        if cell:
                            try:
                                ipaddress.ip_address(cell)
                                ips.append(cell)
                                break
                            except ValueError:
                                continue
        
        print(f"Read {len(ips)} IP addresses from {csv_file}")
        return ips
        
    except FileNotFoundError:
        print(f"Error: CSV file '{csv_file}' not found")
        return []
    except Exception as e:
        print(f"Error reading CSV file: {str(e)}")
        return []

def save_to_excel(results, output_file):
    """Save results to Excel file with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "DNS Lookup Results"
    
    # Headers
    headers = ["IP Address", "Hostname"]
    ws.append(headers)
    
    # Format headers
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
    
    # Add data
    for ip, hostname in results:
        ws.append([ip, hostname])
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save file
    wb.save(output_file)
    print(f"Results saved to {output_file}")

def main():
    if len(sys.argv) != 3:
        print("Usage: python dns_lookup_windows.py <input_csv_file> <output_excel_file>")
        print("Example: python dns_lookup_windows.py ip_addresses.csv dns_results.xlsx")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_excel = sys.argv[2]
    
    system_name = platform.system()
    print(f"Starting DNS lookup process on {system_name}...")
    print("=" * 60)
    
    # Get actual system DNS servers
    system_dns_servers = get_system_dns_servers()
    
    # Read IP addresses from CSV
    ips = read_ips_from_csv(input_csv)
    if not ips:
        print("No valid IP addresses found in the CSV file")
        sys.exit(1)
    
    print(f"Processing {len(ips)} IP addresses...")
    print(f"Using system DNS servers: {system_dns_servers}")
    print("RFC 1918 (private) addresses will use system DNS")
    print("Public addresses will use system DNS first, external DNS as fallback")
    print("Using multiple DNS resolution methods for maximum compatibility")
    print("=" * 60)
    
    start_time = time.time()
    
    # Process IPs in smaller batches for Windows compatibility
    batch_size = 250
    all_results = []
    
    for i in range(0, len(ips), batch_size):
        batch = ips[i:i + batch_size]
        print(f"Processing batch {i//batch_size + 1} ({len(batch)} addresses)...")
        
        batch_results = process_ip_batch(batch, system_dns_servers)
        all_results.extend(batch_results)
        
        print(f"Completed batch {i//batch_size + 1}")
        print("-" * 40)
    
    elapsed_time = time.time() - start_time
    print(f"DNS lookups completed in {elapsed_time:.2f} seconds")
    print(f"Average: {elapsed_time/len(ips):.3f} seconds per IP")
    
    # Save to Excel
    print("Saving results to Excel...")
    save_to_excel(all_results, output_excel)
    
    # Summary
    successful_lookups = sum(1 for _, hostname in all_results 
                           if not hostname.startswith(('DNS error:', 'No PTR record', 'DNS timeout', 'Error:', 'Socket')))
    
    print("=" * 60)
    print("SUMMARY:")
    print(f"Total IP addresses processed: {len(all_results)}")
    print(f"Successful DNS lookups: {successful_lookups}")
    print(f"Failed lookups: {len(all_results) - successful_lookups}")
    print(f"Success rate: {(successful_lookups/len(all_results)*100):.1f}%")
    print(f"Results saved to: {output_excel}")
    print("=" * 60)
    print("System:", system_name)
    print("DNS servers used:", system_dns_servers)

if __name__ == "__main__":
    main()
