#!/usr/bin/env python3
"""
fex_report.py - Cisco NX-OS FEX Interface Report Generator

Connects to NX-OS parent switches, runs 'show interface status' and
'show interface description' for specified FEX ports, and writes the
results to a formatted Excel spreadsheet.

Requirements:  pip install netmiko openpyxl
"""

import re
import logging
from datetime import datetime
from netmiko import ConnectHandler
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from tools import get_netmiko_creds, getScriptName, setupLogging


# ====================================================================
# USER INPUT — edit these two structures before each run
# ====================================================================
# Key   = switch hostname or IP (the parent switch that owns the FEX)
# Value = list of FEX numbers to report on
SWITCHES = {
    "switch1.example.com": [101, 102],
    "switch2.example.com": [103],
}

# Port range and module inside each FEX (adjust for your hardware)
FEX_PORT_RANGE = "1-48"
FEX_MODULE = 1


# ====================================================================
# Standard credentials and logging setup (from tools.py)
# ====================================================================
scriptName = getScriptName()
timestamp = datetime.now().strftime('%Y%m%d_%H%M')
log_path = setupLogging(scriptName, timestamp)
logger = logging.getLogger(__name__)
netmikoUser, passwd, enable = get_netmiko_creds()

OUTPUT_FILE = f"fex_interface_report_{timestamp}.xlsx"


# ====================================================================
# Parsing helpers
# ====================================================================

def _col_positions(header, names):
    """Derive (start, end) slicing positions from a fixed-width header line."""
    positions = []
    for i, name in enumerate(names):
        start = header.index(name)
        end = header.index(names[i + 1]) if i + 1 < len(names) else None
        positions.append((start, end))
    return positions


def _extract(line, positions, keys):
    """Slice one data line according to precomputed column positions."""
    row = {}
    for key, (start, end) in zip(keys, positions):
        row[key] = (line[start:end] if end else line[start:]).strip()
    return row


def parse_interface_status(raw):
    """
    Parse NX-OS 'show interface <range> status'.

    Expected columns (7):
        Port  Name  Status  Vlan  Duplex  Speed  Type
    """
    COL_NAMES = ["Port", "Name", "Status", "Vlan", "Duplex", "Speed", "Type"]
    KEYS      = ["port", "name", "status", "vlan", "duplex", "speed", "type"]
    header = None
    results = []
    past_sep = False

    for line in raw.splitlines():
        if not header and all(c in line for c in ("Port", "Status", "Vlan")):
            header = line
            past_sep = False
            continue
        if header and line.lstrip().startswith("---"):
            past_sep = True
            continue
        if past_sep and line.strip():
            results.append(line)

    if not header:
        return []

    pos = _col_positions(header, COL_NAMES)
    return [_extract(l, pos, KEYS) for l in results]


def parse_interface_description(raw):
    """
    Parse NX-OS 'show interface <range> description'.

    Handles two common formats:
      4-col:  Port  Type  Speed  Description
      2-col:  Interface  Description
    Returns (list-of-dicts, is_four_col).
    """
    header = None
    data_lines = []
    past_sep = False

    for line in raw.splitlines():
        if not header and "Description" in line and ("Port" in line or "Interface" in line):
            header = line
            past_sep = False
            continue
        if header and line.lstrip().startswith("---"):
            past_sep = True
            continue
        if past_sep and line.strip():
            data_lines.append(line)

    if not header:
        return [], True

    four_col = "Type" in header and "Speed" in header

    if four_col:
        col_names = ["Port", "Type", "Speed", "Description"]
        keys      = ["interface", "type", "speed", "description"]
    else:
        col_names = ["Interface", "Description"]
        keys      = ["interface", "description"]

    pos = _col_positions(header, col_names)
    rows = [_extract(l, pos, keys) for l in data_lines]
    return rows, four_col


def normalize_intf(name):
    """Normalize to short form for reliable matching."""
    return re.sub(r"(?i)^ethernet", "Eth", name.strip()).lower()


def intf_sort_key(entry):
    """Sort interfaces numerically by the trailing port number."""
    m = re.search(r"/(\d+)$", entry.get("port", ""))
    return int(m.group(1)) if m else 0


# ====================================================================
# Data collection
# ====================================================================

def collect_switch(host, fex_list):
    """SSH to one switch and return data for every requested FEX."""
    logger.info(f"Connecting to {host}")
    device = {
        "device_type": "cisco_nxos",
        "host": host,
        "username": netmikoUser,
        "password": passwd,
        "secret": enable,
        "fast_cli": False,
        "timeout": 60,
    }

    fex_results = []  # [(fex, status_rows, desc_rows, four_col), …]

    try:
        with ConnectHandler(**device) as conn:
            conn.enable()
            logger.info(f"Connected to {host}")

            for fex in fex_list:
                intf = f"Eth{fex}/{FEX_MODULE}/{FEX_PORT_RANGE}"
                logger.info(f"  FEX {fex}  ({intf})")

                raw_status = conn.send_command(f"show interface {intf} status")
                if "Invalid" in raw_status or "%" in raw_status:
                    logger.warning(f"  status command error on {host} FEX {fex}: "
                                   f"{raw_status.strip().splitlines()[0]}")
                    continue
                status_rows = parse_interface_status(raw_status)

                raw_desc = conn.send_command(f"show interface {intf} description")
                if "Invalid" in raw_desc or "%" in raw_desc:
                    logger.warning(f"  description command error on {host} FEX {fex}: "
                                   f"{raw_desc.strip().splitlines()[0]}")
                    desc_rows, four_col = [], True
                else:
                    desc_rows, four_col = parse_interface_description(raw_desc)

                status_rows.sort(key=intf_sort_key)
                logger.info(f"    {len(status_rows)} status / {len(desc_rows)} description entries")
                fex_results.append((fex, status_rows, desc_rows, four_col))

    except Exception as e:
        logger.error(f"Failed connecting to {host}: {e}")

    return fex_results


# ====================================================================
# Excel report
# ====================================================================

# Column layout (1-based)
#   A=1  Host Switch
#   B=2  FEX
#   C=3  Port
#   D=4  Description  (from show interface description — replaces truncated Name)
#   E=5  Status
#   F=6  Vlan
#   G=7  Duplex
#   H=8  Speed
#   I=9  Type
TOTAL_COLS = 9


def build_report(all_data):
    """Create the formatted .xlsx workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FEX Interface Report"

    # ---- styles ----
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")

    # ---- headers ----
    headers = [
        "Host Switch", "FEX",                               # A-B
        "Port", "Description", "Status", "Vlan",            # C-F
        "Duplex", "Speed", "Type",                          # G-I
    ]
    for ci, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=text)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align

    # ---- data rows ----
    row = 2
    for host, fex_results in all_data:
        for fex, status_rows, desc_rows, four_col in fex_results:

            # Build lookup: normalized interface → description dict
            desc_map = {}
            for d in desc_rows:
                desc_map[normalize_intf(d["interface"])] = d

            for entry in status_rows:
                # Use full description from show int description instead
                # of the truncated Name from show int status
                key = normalize_intf(entry.get("port", ""))
                desc = desc_map.get(key, {})
                full_desc = desc.get("description", entry.get("name", ""))

                ws.cell(row=row, column=1, value=host)
                ws.cell(row=row, column=2, value=fex)
                ws.cell(row=row, column=3, value=entry.get("port", ""))
                ws.cell(row=row, column=4, value=full_desc)
                ws.cell(row=row, column=5, value=entry.get("status", ""))
                ws.cell(row=row, column=6, value=entry.get("vlan", ""))
                ws.cell(row=row, column=7, value=entry.get("duplex", ""))
                ws.cell(row=row, column=8, value=entry.get("speed", ""))
                ws.cell(row=row, column=9, value=entry.get("type", ""))

                # gray background for anything not "connected"
                if entry.get("status", "").lower() != "connected":
                    for c in range(1, TOTAL_COLS + 1):
                        ws.cell(row=row, column=c).fill = gray_fill

                row += 1

    # ---- column widths ----
    for ci in range(1, TOTAL_COLS + 1):
        letter = get_column_letter(ci)
        max_len = max(
            (len(str(cell.value or "")) for cell in ws[letter]),
            default=8,
        )
        ws.column_dimensions[letter].width = max_len + 3

    # ---- freeze top row ----
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    logger.info(f"Report saved to {OUTPUT_FILE}")


# ====================================================================
# Main
# ====================================================================

def main():
    all_data = []  # [(host, fex_results), …]

    for host, fex_list in SWITCHES.items():
        fex_results = collect_switch(host, fex_list)
        all_data.append((host, fex_results))

    if not any(fr for _, fr in all_data):
        logger.warning("No data collected — verify connectivity and FEX numbers.")
        return

    build_report(all_data)
    print(f"Report: {OUTPUT_FILE}")
    print(f"Log:    {log_path}")


if __name__ == "__main__":
    main()
