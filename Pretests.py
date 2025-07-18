import os
from datetime import datetime
from netmiko import ConnectHandler
from openpyxl import Workbook
from openpyxl.styles import Alignment
from getCreds import getNetmikoCreds
import tools

# --- Setup ---
scriptName = tools.getScriptName()
logger = tools.setupLogging(scriptName)

# Per-switch commands
switchCommands = {
    "switch1.example.com": ["show version", "show inventory"],
    "switch2.example.com": ["show vlan brief"],
    "switch3.example.com": ["show cdp neighbors"],
    "switch4.example.com": ["show interface brief"],
}

# Output file with timestamp
timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
outputFile = f"{scriptName}_{timestamp}.xlsx"

def runCommands():
    username, password, enable = getNetmikoCreds()
    wb = Workbook()
    first = True

    for switch, commands in switchCommands.items():
        logger.info(f"Connecting to {switch}")
        device = {
            "device_type": "cisco_nxos",
            "host": switch,
            "username": username,
            "password": password,
            "secret": enable,
        }

        try:
            conn = ConnectHandler(**device)
            conn.enable()
        except Exception as e:
            logger.error(f"Connection failed for {switch}: {e}")
            continue

        # Setup worksheet
        if first:
            ws = wb.active
            ws.title = switch
            first = False
        else:
            ws = wb.create_sheet(title=switch)

        ws.append(["Command", "Output"])
        ws.freeze_panes = "A2"
        wrap_style = Alignment(wrap_text=True)

        for cmd in commands:
            logger.info(f"{switch}: Running '{cmd}'")
            try:
                output = conn.send_command(cmd)
                row = [cmd, output]
            except Exception as e:
                logger.error(f"Command failed: '{cmd}' on {switch}: {e}")
                row = [cmd, f"ERROR: {e}"]

            ws.append(row)
            ws.cell(row=ws.max_row, column=2).alignment = wrap_style

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 100)

        conn.disconnect()

    wb.save(outputFile)
    logger.info(f"Results saved to {outputFile}")

if __name__ == "__main__":
    runCommands()


import os
from datetime import datetime
from netmiko import ConnectHandler
from openpyxl import Workbook
from openpyxl.styles import Alignment
from getCreds import get_netmiko_creds
import tools

# --- Setup ---
scriptName = tools.getScriptName()
logger = tools.setupLogging(scriptName)

# Per-switch command list
switchCommands = {
    "switch1.example.com": ["show version", "show inventory"],
    "switch2.example.com": ["show vlan brief"],
    "switch3.example.com": ["show cdp neighbors"],
    "switch4.example.com": ["show interface brief"],
}

def main():
    username, password, enable = get_netmiko_creds()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    outputFile = f"{scriptName}_{timestamp}.xlsx"
    wb = Workbook()
    firstSheet = True
    wrap = Alignment(wrap_text=True)

    for switch, commands in switchCommands.items():
        logger.info(f"Connecting to {switch}")
        device = {
            "device_type": "cisco_nxos",
            "host": switch,
            "username": username,
            "password": password,
            "secret": enable,
        }

        try:
            conn = ConnectHandler(**device)
            conn.enable()
        except Exception as e:
            logger.error(f"Connection failed to {switch}: {e}")
            continue

        ws = wb.active if firstSheet else wb.create_sheet()
        ws.title = switch
        firstSheet = False

        ws.append(["Command", "Output"])
        ws.freeze_panes = "A2"

        for cmd in commands:
            logger.info(f"{switch}: Running '{cmd}'")
            try:
                output = conn.send_command(cmd)
                ws.append([cmd, output])
                ws.cell(row=ws.max_row, column=2).alignment = wrap
            except Exception as e:
                logger.error(f"Command failed on {switch}: {cmd}: {e}")
                ws.append([cmd, f"ERROR: {e}"])

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 2, 100)

        conn.disconnect()

    wb.save(outputFile)
    logger.info(f"Results saved to {outputFile}")

if __name__ == "__main__":
    main()
