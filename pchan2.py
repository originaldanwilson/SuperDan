import os
import re
from netmiko import ConnectHandler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tools # Import your tools.py

# --- Configuration ---
# List of DNS names or IP addresses for your devices
# Replace with your actual device list
DEVICES = [
    "nxos-device1.yourdomain.com",
    "nxos-device2.yourdomain.com",
    "some-other-device.yourdomain.com", # This one will be skipped if not NX-OS
    "another-nxos.yourdomain.com"
]

OUTPUT_FILENAME = "nxos_cdp_port_channel_correlation.xlsx"

# --- Setup Logging ---
logger = tools.get_logger()

# --- Main Script Functions ---
def get_device_info(device_name, username, password):
    """
    Connects to a device, checks if it's NX-OS, and gathers CDP neighbors and port-channel summary.
    Returns (hostname, cdp_output, port_channel_output) or None if not NX-OS or connection fails.
    """
    device_info = {
        'device_type': 'cisco_nxos', # Assuming we're targeting NX-OS
        'host': device_name,
        'username': username,
        'password': password,
        'secret': password, # Often the same for simple setups, or prompt for enable password
        'fast_cli': False, # Set to True for faster execution if needed
    }
    net_connect = None
    hostname = None
    cdp_output = ""
    port_channel_output = ""
    is_nxos = False

    try:
        logger.info(f"Attempting to connect to {device_name}...")
        net_connect = ConnectHandler(**device_info)
        hostname = net_connect.find_prompt().replace('#', '').replace('>', '').strip() # Get actual hostname

        # Check if it's an NX-OS device
        show_version = net_connect.send_command("show version | include NX-OS", use_textfsm=False)
        if "NX-OS" in show_version:
            is_nxos = True
            logger.info(f"Successfully connected to {hostname} ({device_name}). It's an NX-OS device.")
            cdp_output = net_connect.send_command("show cdp neighbors detail", use_textfsm=False)
            port_channel_output = net_connect.send_command("show port-channel summary", use_textfsm=False)
        else:
            logger.warning(f"{device_name} is not identified as an NX-OS device. Skipping.")
            return None, None, None, None

    except Exception as e:
        logger.error(f"Failed to connect or retrieve info from {device_name}: {e}")
        return None, None, None, None
    finally:
        if net_connect:
            net_connect.disconnect()

    return hostname, cdp_output, port_channel_output, is_nxos

def parse_cdp_neighbors(cdp_output):
    """
    Parses 'show cdp neighbors detail' output to extract local interface, neighbor device, and neighbor interface.
    Returns a list of dictionaries.
    """
    cdp_entries = []
    # Regex to find blocks for each neighbor
    neighbor_blocks = re.split(r'----------------------------------------', cdp_output)

    for block in neighbor_blocks:
        if "Device ID:" in block:
            local_interface_match = re.search(r"Interface:\s*(\S+),\s*Port ID", block)
            device_id_match = re.search(r"Device ID:\s*(\S+)", block)
            neighbor_interface_match = re.search(r"Port ID \(outgoing port\):\s*(\S+)", block)

            if local_interface_match and device_id_match and neighbor_interface_match:
                cdp_entries.append({
                    "local_interface": local_interface_match.group(1),
                    "neighbor_device": device_id_match.group(1),
                    "neighbor_interface": neighbor_interface_match.group(1)
                })
    return cdp_entries

def parse_port_channel_summary(port_channel_output):
    """
    Parses 'show port-channel summary' output to extract port-channel ID and its member interfaces.
    Returns a dictionary mapping member interfaces to port-channel IDs.
    Handles multi-line member interface listings.
    """
    port_channel_map = {}
    lines = port_channel_output.splitlines()

    current_po_id = None
    for line in lines:
        # Skip header and empty lines
        if not line.strip() or "Group" in line or "---" in line or "Flags:" in line:
            continue

        # Pattern for the start of a new port-channel entry
        # Example: 1   Po1(SU)  Eth  LACP      Eth1/1(P)    Eth1/2(P)
        # Note: Added an optional group number at the start, as seen in some outputs
        po_start_match = re.match(r"^\s*(\d*\s*)?(Po\d+)\(\S+\)\s+\S+\s+(?:LACP|NONE|PAGP)\s*(.*)", line)

        if po_start_match:
            # Found a new port-channel entry
            group_num, po_id, interfaces_str = po_start_match.groups()
            current_po_id = po_id.strip()

            # Find interfaces on this line
            member_interfaces = re.findall(r"(Eth\d+/\d+)\(P\)", interfaces_str)
            for member_if in member_interfaces:
                port_channel_map[member_if] = current_po_id
        elif current_po_id:
            # This line might contain additional member interfaces for the current port-channel
            # It should ideally be indented or not match the start pattern for a new PO
            additional_interfaces = re.findall(r"(Eth\d+/\d+)\(P\)", line)
            for member_if in additional_interfaces:
                port_channel_map[member_if] = current_po_id
        # If current_po_id is None and it's not a new PO line, it's likely a blank line or
        # an uninteresting line between PO entries, so we ignore it.
    return port_channel_map


def write_to_excel(data, filename):
    """
    Writes the correlated data to an Excel spreadsheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CDP-PortChannel Correlation"

    # Define headers
    headers = [
        "Hostname",
        "Local Interface",
        "Port-Channel ID",
        "Neighbor Device",
        "Neighbor Interface"
    ]
    ws.append(headers)

    # Apply bold and fill to header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Write data rows
    for row_data in data:
        ws.append([
            row_data.get("hostname", ""),
            row_data.get("local_interface", ""),
            row_data.get("port_channel_id", ""),
            row_data.get("neighbor_device", ""),
            row_data.get("neighbor_interface", "")
        ])

    # Auto-size columns (basic attempt, adjust as needed)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)
    logger.info(f"Data successfully written to {os.path.abspath(filename)}")

def main():
    username, password = tools.get_netmiko_credentials()
    all_correlated_data = []

    for device_name in DEVICES:
        hostname, cdp_output, port_channel_output, is_nxos = get_device_info(device_name, username, password)

        if hostname and is_nxos:
            cdp_entries = parse_cdp_neighbors(cdp_output)
            port_channel_map = parse_port_channel_summary(port_channel_output)

            # Correlate CDP neighbors with port-channel interfaces
            for cdp_entry in cdp_entries:
                local_if = cdp_entry["local_interface"]
                correlated_entry = {
                    "hostname": hostname,
                    "local_interface": local_if,
                    "neighbor_device": cdp_entry["neighbor_device"],
                    "neighbor_interface": cdp_entry["neighbor_interface"]
                }

                # Check if the local interface is part of a port-channel
                if local_if in port_channel_map:
                    correlated_entry["port_channel_id"] = port_channel_map[local_if]
                else:
                    correlated_entry["port_channel_id"] = "N/A (Standalone)"
                all_correlated_data.append(correlated_entry)

    if all_correlated_data:
        write_to_excel(all_correlated_data, OUTPUT_FILENAME)
    else:
        logger.warning("No NX-OS device data was collected to write to Excel.")

if __name__ == "__main__":
    main()
