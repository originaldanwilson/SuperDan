#!/usr/bin/env python3
"""
Ping network devices loaded from a JSON inventory file.
Pings each device by both Display Name (DNS) and IP Address concurrently,
then writes results to a comma-delimited CSV file.
"""

import json
import socket
import subprocess
import sys
import os
from dataclasses import dataclass, field
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


@dataclass
class PingResult:
    """Stores the outcome of a single ping attempt."""
    target: str                # what was pinged (hostname or IP)
    target_type: str           # "DisplayName" or "IP_Address"
    resolved_ip: str = "N/A"   # actual IP the target resolved to
    reachable: bool = False
    packet_loss: str = "100%"
    rtt_min: str = "N/A"
    rtt_avg: str = "N/A"
    rtt_max: str = "N/A"
    error: Optional[str] = None


@dataclass
class Device:
    """Represents a network device from the JSON inventory."""
    display_name: str
    ip_address: str
    machine_type: str
    ios_image: str
    ping_by_name: Optional[PingResult] = field(default=None, repr=False)
    ping_by_ip: Optional[PingResult] = field(default=None, repr=False)

    # Map our fields to possible JSON key variations (checked case-insensitively)
    _KEY_MAP = {
        "display_name": ["display name", "displayname", "caption", "nodename", "node_name", "hostname"],
        "ip_address":   ["ip_address", "ipaddress", "ip address", "ip", "managementip"],
        "machine_type": ["machinetype", "machine_type", "machine type", "devicetype", "vendor", "model", "sysobjectid"],
        "ios_image":    ["iosimage", "ios_image", "ios image", "software", "osversion", "firmware"],
    }

    @staticmethod
    def _find_key(d: dict, candidates: list[str]) -> str | None:
        """Find the first matching key in d (case-insensitive)."""
        lower_map = {k.lower(): k for k in d}
        for candidate in candidates:
            if candidate in lower_map:
                return lower_map[candidate]
        return None

    @classmethod
    def from_dict(cls, d: dict) -> "Device":
        resolved = {}
        missing = []
        for field_name, candidates in cls._KEY_MAP.items():
            key = cls._find_key(d, candidates)
            if key is not None:
                resolved[field_name] = d[key]
            else:
                missing.append(field_name)

        if missing:
            print(f"\n  WARNING: Could not map fields: {missing}")
            print(f"  Available keys in JSON: {list(d.keys())}")
            # Fill missing fields with 'UNKNOWN' so it still runs
            for m in missing:
                resolved[m] = "UNKNOWN"

        return cls(**resolved)


def ping(target: str, target_type: str, count: int = 2, timeout: int = 2) -> PingResult:
    """
    Ping a target on Linux and parse the output.
    Uses -c for count and -W for per-packet timeout (seconds).
    """
    result = PingResult(target=target, target_type=target_type)

    # Resolve the target to an actual IP address
    try:
        result.resolved_ip = socket.gethostbyname(target)
    except socket.gaierror:
        result.resolved_ip = "DNS_FAIL"

    try:
        proc = subprocess.run(
            ["ping", "-c", str(count), "-W", str(timeout), target],
            capture_output=True,
            text=True,
            timeout=count * timeout + 5,  # generous overall timeout
        )
        output = proc.stdout

        # Parse packet loss  (e.g. "0% packet loss")
        for line in output.splitlines():
            if "packet loss" in line:
                for part in line.split(","):
                    part = part.strip()
                    if "packet loss" in part:
                        result.packet_loss = part.split()[0]
                        break

        # Parse RTT  (e.g. "rtt min/avg/max/mdev = 0.031/0.035/0.040/0.004 ms")
        for line in output.splitlines():
            if "rtt" in line or "round-trip" in line:
                stats = line.split("=")[1].strip().split("/")
                result.rtt_min = stats[0].strip()
                result.rtt_avg = stats[1].strip()
                result.rtt_max = stats[2].strip()
                break

        result.reachable = proc.returncode == 0

    except subprocess.TimeoutExpired:
        result.error = "Ping timed out"
    except FileNotFoundError:
        result.error = "ping command not found"
    except Exception as e:
        result.error = str(e)

    return result


def ping_device(device: Device) -> Device:
    """Ping a device by both display name and IP address."""
    device.ping_by_name = ping(device.display_name, "DisplayName")
    device.ping_by_ip = ping(device.ip_address, "IP_Address")
    return device


def load_devices(json_path: str) -> list[Device]:
    """Load device list from a JSON file.

    Handles three common formats:
      1. A bare JSON array:   [{...}, {...}, ...]
      2. An object with one key whose value is the array:
         {"results": [{...}, {...}, ...]}
      3. An object with multiple keys — looks for the first
         key whose value is a list of dicts.
    """
    with open(json_path, "r") as f:
        data = json.load(f)

    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        # Find the first value that is a list of dicts
        items = None
        for key, val in data.items():
            if isinstance(val, list) and val and isinstance(val[0], dict):
                items = val
                print(f"  Using key '{key}' from JSON ({len(val)} entries)")
                break
        if items is None:
            sys.exit(f"ERROR: Could not find a device list in {json_path}. "
                     f"Top-level keys: {list(data.keys())}")
    else:
        sys.exit(f"ERROR: Unexpected JSON type: {type(data).__name__}")

    return [Device.from_dict(entry) for entry in items]


def write_xlsx(devices: list[Device], output_path: str) -> None:
    """Write ping results to an Excel spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ping Results"

    headers = [
        "Display Name", "IP Address (from JSON)", "Machine Type", "IOS Image",
        "Ping Target", "Ping Target Type", "Resolved IP",
        "Reachable", "Packet Loss", "RTT Min (ms)", "RTT Avg (ms)", "RTT Max (ms)",
        "Error", "Timestamp",
    ]

    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_num = 2

    for dev in devices:
        for pr in (dev.ping_by_name, dev.ping_by_ip):
            if pr is None:
                continue
            values = [
                dev.display_name,
                dev.ip_address,
                dev.machine_type,
                dev.ios_image,
                pr.target,
                pr.target_type,
                pr.resolved_ip,
                "Yes" if pr.reachable else "No",
                pr.packet_loss,
                pr.rtt_min,
                pr.rtt_avg,
                pr.rtt_max,
                pr.error or "",
                timestamp,
            ]
            for col_num, val in enumerate(values, 1):
                ws.cell(row=row_num, column=col_num, value=val)
            row_num += 1

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(output_path)


def main():
    json_path = sys.argv[1] if len(sys.argv) > 1 else "devices.json"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "ping_results.xlsx"

    if not os.path.isfile(json_path):
        print(f"ERROR: File not found: {json_path}")
        sys.exit(1)

    devices = load_devices(json_path)
    print(f"Loaded {len(devices)} devices from {json_path}")

    # Ping all devices concurrently — one thread per device
    max_workers = min(20, len(devices))  # cap threads to avoid flooding
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(ping_device, dev): dev for dev in devices}
        for future in as_completed(futures):
            dev = future.result()
            name_status = "UP" if dev.ping_by_name and dev.ping_by_name.reachable else "DOWN"
            ip_status = "UP" if dev.ping_by_ip and dev.ping_by_ip.reachable else "DOWN"
            print(f"  {dev.display_name:<35} name={name_status:<4}  ip={ip_status:<4}  ({dev.machine_type})")

    write_xlsx(devices, output_path)
    print(f"\nResults saved to {output_path}")


if __name__ == "__main__":
    main()
