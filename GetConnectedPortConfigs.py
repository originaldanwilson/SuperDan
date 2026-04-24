#!/usr/bin/env python3
"""
GetConnectedPortConfigs.py

Query one or more Cisco NX-OS switches (including any attached FEXs) and
produce an Excel workbook that, for each switch, lists:

    * every interface whose status is "connected"
    * the speed and duplex of that connection
    * the full running-config block for that interface (listed right
      after the port in the same row)

One worksheet is created per switch so it is easy to copy / compare the
connected-port list to a second spreadsheet that maps the old port
numbers to the new switch's port numbers.  A final "summary" sheet lists
the per-switch connected-port counts.

The resulting list of connected ports (per switch) is also written to a
plain-text file so it can be sourced into a follow-up script that
correlates old->new interface names and builds the configuration that
will be pushed onto the new switch.

Modeled after pdesc34.py / ExploreFex.py / postCheck.py.
"""

import logging
import os
import re
from collections import defaultdict
from datetime import datetime

from netmiko import ConnectHandler
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from tools import get_netmiko_creds, getScriptName, setupLoggingNew


# ---------------------------------------------------------------------------
# EDIT THIS LIST - the switches to query
# ---------------------------------------------------------------------------
SWITCHES = [
    "nxos-switch-01",
    # "nxos-switch-02",
    # "nxos-switch-03",
]

# Interface-name prefixes that should never be considered "connected ports"
# for the purpose of a server migration, even if they happen to show up as
# connected in "show interface status".
EXCLUDE_PREFIXES = ("mgmt", "lo", "vlan", "nve", "po")

# Excel column headers (kept in this order in every per-switch sheet)
HEADERS = [
    "Switch",
    "Interface",
    "Status",
    "VLAN",
    "Description",
    "Speed",
    "Duplex",
    "Type",
    "Configuration",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def parse_running_config(run_cfg_text):
    """
    Parse the output of `show running-config` into a dict keyed by
    interface name (exactly as it appears after the `interface ` token)
    whose value is the raw config block for that interface, including
    the `interface <name>` header line.

    Works for NX-OS native slot/port (Eth1/1) and FEX three-tuple
    (Eth101/1/1) names, as well as port-channels and mgmt interfaces.
    """
    configs = {}
    current_name = None
    current_lines = []

    iface_re = re.compile(r"^interface\s+(\S+)\s*$", re.IGNORECASE)

    for line in run_cfg_text.splitlines():
        m = iface_re.match(line)
        if m:
            # flush previous
            if current_name is not None:
                configs[current_name] = "\n".join(current_lines).rstrip()
            current_name = m.group(1)
            current_lines = [line.rstrip()]
            continue

        if current_name is None:
            continue

        # Interface block ends at the first non-indented, non-empty line.
        if line and not line.startswith((" ", "\t")):
            configs[current_name] = "\n".join(current_lines).rstrip()
            current_name = None
            current_lines = []
            continue

        current_lines.append(line.rstrip())

    if current_name is not None:
        configs[current_name] = "\n".join(current_lines).rstrip()

    return configs


def is_excluded(interface_name):
    name = interface_name.lower()
    return name.startswith(EXCLUDE_PREFIXES)


def init_sheet(wb, title):
    """
    Create a worksheet with standard headers.  Worksheet titles are
    capped at 31 chars and sanitized of characters Excel does not allow.
    """
    safe = re.sub(r"[\\/*?:\[\]]", "_", str(title))[:31] or "switch"
    ws = wb.create_sheet(title=safe)
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    return ws


def auto_width_all(wb, config_col_letter="I", config_col_width=80):
    """
    Auto-size every column.  The Configuration column is wide and
    wrapped so the multi-line running-config is readable.
    """
    for ws in wb.worksheets:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            if letter == config_col_letter:
                ws.column_dimensions[letter].width = config_col_width
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0)
                    for c in col
                )
                ws.column_dimensions[letter].width = min(max_len + 2, 40)


# ---------------------------------------------------------------------------
# Per-switch collection
# ---------------------------------------------------------------------------
def collect_switch(hostname, netmikoUser, passwd, enable, wb,
                   summary_tracker, ports_by_switch):
    """
    Connect to `hostname`, find every connected interface, grab its
    running-config and write everything onto a worksheet named after the
    switch.  Updates `summary_tracker` with the per-FEX counts and fills
    `ports_by_switch[hostname]` with the list of connected interface
    names (used to build the plain-text port list).
    """
    device = {
        "device_type": "cisco_nxos",
        "host": hostname,
        "username": netmikoUser,
        "password": passwd,
        "secret": enable,
        "fast_cli": False,
        "timeout": 60,
    }

    try:
        logging.info(f"Connecting to {hostname}")
        with ConnectHandler(**device) as conn:
            conn.enable()
            conn.send_command("terminal length 0")

            status_rows = conn.send_command(
                "show interface status", use_textfsm=True, read_timeout=120
            )
            run_cfg_text = conn.send_command(
                "show running-config", read_timeout=300
            )
    except Exception as e:
        logging.error(f"Failed to query {hostname}: {e}")
        return

    if not isinstance(status_rows, list):
        logging.error(
            f"{hostname}: TextFSM parse of 'show interface status' failed; "
            "skipping.  Make sure ntc-templates is installed."
        )
        return

    configs = parse_running_config(run_cfg_text)
    ws = init_sheet(wb, hostname)

    connected = 0
    for entry in status_rows:
        status = (entry.get("status") or "").lower()
        if status != "connected":
            continue

        intf = entry.get("port", "")
        if not intf or is_excluded(intf):
            continue

        desc = entry.get("name", "")
        vlan = entry.get("vlan", "")
        speed = entry.get("speed", "")
        duplex = entry.get("duplex", "")
        iftype = entry.get("type", "")

        # Try a few sensible key variants so we survive "Eth1/1" vs
        # "Ethernet1/1" mismatches between `show interface status` and
        # `show running-config`.
        cfg = (
            configs.get(intf)
            or configs.get(intf.replace("Eth", "Ethernet"))
            or configs.get(intf.replace("Ethernet", "Eth"))
            or ""
        )

        ws.append([
            hostname, intf, status, vlan, desc, speed, duplex, iftype, cfg,
        ])
        connected += 1

        # Track by FEX id (first numeric slot >=100 is a FEX, else native)
        m = re.match(r"Eth(?:ernet)?(\d+)/", intf)
        if m:
            slot = int(m.group(1))
            label = str(slot) if slot >= 100 else "native"
        else:
            label = "other"
        summary_tracker[hostname][label] += 1

        ports_by_switch[hostname].append(intf)

    logging.info(f"{hostname}: {connected} connected ports written")


# ---------------------------------------------------------------------------
# Summary & port-list outputs
# ---------------------------------------------------------------------------
def write_summary_sheet(wb, summary_tracker):
    ws = wb.create_sheet(title="summary", index=0)
    all_labels = sorted(
        {lbl for per in summary_tracker.values() for lbl in per.keys()},
        key=lambda x: (x != "native", x),
    )
    header = ["Switch", "Total Connected"] + all_labels
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for hostname in sorted(summary_tracker):
        per = summary_tracker[hostname]
        total = sum(per.values())
        row = [hostname, total] + [per.get(lbl, 0) for lbl in all_labels]
        ws.append(row)

    ws.freeze_panes = "B2"


def write_port_list_file(ports_by_switch, path):
    """
    Write a simple text file of the connected ports, grouped per switch.
    This is the list you will source in the follow-up script that maps
    old port -> new port on the destination switch.

    Format:

        # <hostname>
        CONNECTED_PORTS["<hostname>"] = [
            "Eth1/1",
            "Eth1/2",
            ...
        ]
    """
    with open(path, "w", newline="") as f:
        f.write("# Auto-generated by GetConnectedPortConfigs.py\n")
        f.write("# Paste this into your migration script.\n\n")
        f.write("CONNECTED_PORTS = {}\n\n")
        for hostname in sorted(ports_by_switch):
            ports = ports_by_switch[hostname]
            f.write(f"# {hostname} - {len(ports)} connected ports\n")
            f.write(f'CONNECTED_PORTS["{hostname}"] = [\n')
            for p in ports:
                f.write(f'    "{p}",\n')
            f.write("]\n\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    script_name = getScriptName()
    setupLoggingNew(log_file=f"{script_name}_{timestamp}.log")
    logging.info(f"{script_name} started")

    netmikoUser, passwd, enable = get_netmiko_creds()

    wb = Workbook()
    wb.remove(wb.active)  # drop default blank sheet

    summary_tracker = defaultdict(lambda: defaultdict(int))
    ports_by_switch = defaultdict(list)

    for hostname in SWITCHES:
        collect_switch(
            hostname, netmikoUser, passwd, enable,
            wb, summary_tracker, ports_by_switch,
        )

    write_summary_sheet(wb, summary_tracker)
    auto_width_all(wb)

    xlsx_path = f"{script_name}_{timestamp}.xlsx"
    wb.save(xlsx_path)
    logging.info(f"Saved spreadsheet: {xlsx_path}")

    ports_path = f"{script_name}_ports_{timestamp}.txt"
    write_port_list_file(ports_by_switch, ports_path)
    logging.info(f"Saved port list: {ports_path}")

    print(f"Spreadsheet: {os.path.abspath(xlsx_path)}")
    print(f"Port list:   {os.path.abspath(ports_path)}")


if __name__ == "__main__":
    main()
