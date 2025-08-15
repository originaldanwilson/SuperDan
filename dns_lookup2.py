#!/usr/bin/env python3
"""
DNS Lookup Script for IP Addresses
Handles both internal (RFC 1918) and external IP addresses
Outputs results to Excel with proper formatting
"""

import csv
import socket
import ipaddress
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import dns.resolver
import dns.reversename
import time

def is_rfc1918(ip_str):
    """Check if IP address is RFC 1918 (private/internal)"""
    try:
        ip = ipaddress.ip_address(ip_str)
        return ip.is_private
    except ValueError:
        return False

def dns_lookup(ip_address, use_internal=True, retry_count=2):
    """
    Perform reverse DNS lookup on an IP address with retry logic
    
    Args:
        ip_address (str): IP address to lookup
        use_internal (bool): Whether to use internal DNS or external (8.8.8.8)
        retry_count (int): Number of retries for failed lookups
    
    Returns:
        str: Hostname or error message
    """
    last_error = None
    
    for attempt in range(retry_count + 1):
        try:
            # Create a new resolver for each lookup
            resolver = dns.resolver.Resolver()
            
            if use_internal:
                # Use system default DNS servers for internal addresses
                resolver.nameservers = dns.resolver.Resolver().nameservers
                # Add additional common internal DNS servers as fallback
                if attempt > 0:
                    resolver.nameservers.extend(['192.168.1.1', '10.0.0.1'])
            else:
                # Use multiple external DNS servers for better reliability
                if attempt == 0:
                    resolver.nameservers = ['8.8.8.8', '8.8.4.4']
                elif attempt == 1:
                    resolver.nameservers = ['1.1.1.1', '1.0.0.1']  # Cloudflare DNS
                else:
                    resolver.nameservers = ['208.67.222.222', '208.67.220.220']  # OpenDNS
            
            # Set longer timeouts for better reliability
            resolver.timeout = 15  # Individual query timeout
            resolver.lifetime = 30  # Total lifetime for all queries
            
            # Create reverse DNS name
            reverse_name = dns.reversename.from_address(ip_address)
            
            # Perform the lookup
            answer = resolver.resolve(reverse_name, 'PTR')
            hostname = str(answer[0]).rstrip('.')
            
            return hostname
            
        except dns.resolver.NXDOMAIN:
            return "No PTR record found"
        except dns.resolver.Timeout as e:
            last_error = f"DNS timeout (attempt {attempt + 1})"
            if attempt < retry_count:
                time.sleep(1)  # Brief pause before retry
                continue
        except dns.resolver.NoNameservers as e:
            last_error = f"No nameservers available (attempt {attempt + 1})"
            if attempt < retry_count:
                time.sleep(1)
                continue
        except Exception as e:
            last_error = f"DNS error: {str(e)} (attempt {attempt + 1})"
            if attempt < retry_count:
                time.sleep(1)
                continue
    
    return last_error if last_error else "DNS lookup failed"

def process_ip_batch(ip_list):
    """Process a batch of IP addresses with DNS lookups"""
    results = []
    
    with ThreadPoolExecutor(max_workers=20) as executor:
        # Submit all tasks
        future_to_ip = {}
        for ip in ip_list:
            ip = ip.strip()
            if not ip:
                continue
                
            # Determine if we should use internal or external DNS
            use_internal = is_rfc1918(ip)
            future = executor.submit(dns_lookup, ip, use_internal)
            future_to_ip[future] = ip
        
        # Collect results as they complete
        for future in as_completed(future_to_ip):
            ip = future_to_ip[future]
            try:
                hostname = future.result()
                results.append((ip, hostname))
                print(f"Processed {ip}: {hostname}")
            except Exception as e:
                results.append((ip, f"Error: {str(e)}"))
                print(f"Error processing {ip}: {str(e)}")
    
    return results

def read_ips_from_csv(csv_file):
    """Read IP addresses from CSV file"""
    ips = []
    try:
        with open(csv_file, 'r', newline='', encoding='utf-8') as file:
            # Try to detect if first row is header
            sample = file.read(1024)
            file.seek(0)
            sniffer = csv.Sniffer()
            has_header = sniffer.has_header(sample)
            
            reader = csv.reader(file)
            
            if has_header:
                next(reader)  # Skip header row
            
            for row in reader:
                if row:  # Skip empty rows
                    # Take the first column that looks like an IP
                    for cell in row:
                        cell = cell.strip()
                        if cell:
                            try:
                                ipaddress.ip_address(cell)
                                ips.append(cell)
                                break
                            except ValueError:
                                continue
        
        print(f"Read {len(ips)} IP addresses from {csv_file}")
        return ips
        
    except FileNotFoundError:
        print(f"Error: CSV file '{csv_file}' not found")
        return []
    except Exception as e:
        print(f"Error reading CSV file: {str(e)}")
        return []

def save_to_excel(results, output_file):
    """Save results to Excel file with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "DNS Lookup Results"
    
    # Headers
    headers = ["IP Address", "Hostname"]
    ws.append(headers)
    
    # Format headers
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
    
    # Add data
    for ip, hostname in results:
        ws.append([ip, hostname])
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save file
    wb.save(output_file)
    print(f"Results saved to {output_file}")

def main():
    if len(sys.argv) != 3:
        print("Usage: python3 dns_lookup.py <input_csv_file> <output_excel_file>")
        print("Example: python3 dns_lookup.py ip_addresses.csv dns_results.xlsx")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_excel = sys.argv[2]
    
    print("Starting DNS lookup process...")
    print("=" * 50)
    
    # Read IP addresses from CSV
    ips = read_ips_from_csv(input_csv)
    if not ips:
        print("No valid IP addresses found in the CSV file")
        sys.exit(1)
    
    print(f"Processing {len(ips)} IP addresses...")
    print("RFC 1918 (private) addresses will use internal DNS")
    print("Public addresses will use Google DNS (8.8.8.8)")
    print("=" * 50)
    
    start_time = time.time()
    
    # Process IPs in batches to manage memory and connections
    batch_size = 1000
    all_results = []
    
    for i in range(0, len(ips), batch_size):
        batch = ips[i:i + batch_size]
        print(f"Processing batch {i//batch_size + 1} ({len(batch)} addresses)...")
        
        batch_results = process_ip_batch(batch)
        all_results.extend(batch_results)
        
        print(f"Completed batch {i//batch_size + 1}")
        print("-" * 30)
    
    elapsed_time = time.time() - start_time
    print(f"DNS lookups completed in {elapsed_time:.2f} seconds")
    print(f"Average: {elapsed_time/len(ips):.3f} seconds per IP")
    
    # Save to Excel
    print("Saving results to Excel...")
    save_to_excel(all_results, output_excel)
    
    # Summary
    successful_lookups = sum(1 for _, hostname in all_results 
                           if not hostname.startswith(('DNS error:', 'No PTR record', 'DNS timeout')))
    
    print("=" * 50)
    print("SUMMARY:")
    print(f"Total IP addresses processed: {len(all_results)}")
    print(f"Successful DNS lookups: {successful_lookups}")
    print(f"Failed lookups: {len(all_results) - successful_lookups}")
    print(f"Success rate: {(successful_lookups/len(all_results)*100):.1f}%")
    print(f"Results saved to: {output_excel}")

if __name__ == "__main__":
    main()
